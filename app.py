from absensiMethod import (
    cek_tanggal_kerja,
    format_date,
    format_time,
    signInPayload,
    string_to_uuid_like,
    unhadir_absensi,
    uuid_like_to_string,
    get_time_zone_now,
    is_valid_datetime_format,
    upload_to_imgbb,
)
from convert import convert_to_excel, PDF

from flask import (
    Flask,
    redirect,
    url_for,
    render_template,
    request,
    jsonify,
    send_file,
    make_response,
)
from flask_wtf.csrf import CSRFProtect
from pymongo.mongo_client import MongoClient
from dotenv import load_dotenv
import os, hashlib, jwt, datetime
from bson import ObjectId

# from apscheduler.schedulers.background import BackgroundScheduler
from io import BytesIO
from openpyxl import load_workbook

# from werkzeug.utils import secure_filename
from werkzeug.datastructures import FileStorage
from generate_otp import (
    FaqGmailSender,
    OtpPasswordGenerator,
    replyGmailSender,
    TaskGmailNotif,
)
from markupsafe import Markup
import certifi
from flasgger import Swagger

# buat request post sesuai datetime

# loading environment variables from.env file
load_dotenv()
url = os.getenv("MONGODB_URL")
secretKey = os.getenv("SECRET_KEY")

# connecting to mongodb
client = MongoClient(url, tls=True, tlsCAFile=certifi.where())
db = client["absen"]

# inisiasi app
app = Flask(__name__)
# melakukan config
app.config["SECRET_KEY"] = secretKey
# buat csrfprotect
csrf = CSRFProtect(app)


# Tentukan folder untuk menyimpan gambar
upload_folder = os.path.join("/tmp", "img", "user")
# Cek apakah path dapat dituli  s
os.makedirs(upload_folder, exist_ok=True)

app.config["UPLOAD_FOLDER"] = upload_folder

# Daftarkan filter ke Jinja
app.jinja_env.filters["format_date"] = format_date
app.jinja_env.filters["format_time"] = format_time

# url imgbb
imgbb_api_key = os.getenv("IMGBB_API_KEY")

# Ambil data license dari MongoDB
license_name = "Server Side Public License (SSPL) MongoDB"
license_url = "https://www.mongodb.com/licensing/server-side-public-license"
# Konfigurasi Swagger
swagger_template = {
    "swagger": "2.0",
    "info": {
        "title": "API AbsensiKu Winnicode",
        "description": "Dokumentasi API untuk aplikasi absensi Winnicode",
        "version": "1.0.0",
        "termsOfService": "https://www.mongodb.com/legal/terms-and-conditions/cloud#:~:text=MongoDB%20Cloud%20Terms%20of%20Service%201%201.%20Cloud,...%208%208.%20No%20Warranty.%20...%20More%20items?msockid=0963a9adaf8f6eaa1fc1bd71ae8e6f69",
        "contact": {
            "email": "salahudinkoliq10@gmail.com",
        },
        "license": {
            "name": license_name,
            "url": license_url,
        },
    },
    "host": "",
    "basePath": "/",
    "externalDocs": {
        "description": "Find out more about Swagger",
        "url": "http://swagger.io",
    },
    "schemes": ["http", "https"],
    "tags": [
        {
            "name": "Admin / Sub Admin",
            "description": "all endpoint merupakan halaman / proses yang dilakukan di bagian Admin / Sub Admin",
        },
        {
            "name": "Magang / Karyawan",
            "description": "all endpoint merupakan halaman / proses yang dilakukan di bagian Magang / Karyawan",
        },
        {
            "name": "Autentikasi",
            "description": "all endpoint merupakan bagian Autentikasi yaitu sebelum halaman utaama AbsensiKu",
        },
        {
            "name": "Bantuan",
            "description": "all endpoint merupakan bagian Bantuan yaitu Pertanyaan user tiap role dalam mengalami kendala",
        },
        {
            "name": "Landing Page",
            "description": "berisi endpoint pembuka halaman pertama",
        },
        {
            "name": "Error",
            "description": "berisi endpoint mengaami error page atau server",
        },
        {
            "name": "Export",
            "description": "berisi endpoint melakukan export data ke pdf,word, dst",
        },
        {
            "name": "Dokumentasi",
            "description": "berisi endpoint tata cara penggunaan website",
        },
    ],
    "paths": {
        "/": {
            "get": {
                "tags": ["Landing Page"],
                "summary": "Halaman Beranda",
                "description": "Menampilkan halaman beranda dari aplikasi.",
                "responses": {"200": {"description": "Halaman berhasil dimuat."}},
            }
        },
        "/manual/{path}": {
            "get": {
                "tags": ["Dokumentasi"],
                "summary": "Download Manual Book",
                "description": "Endpoint ini digunakan untuk mengunduh manual book berdasarkan path yang diberikan.",
                "parameters": [
                    {
                        "name": "path",
                        "in": "path",
                        "required": True,
                        "type": "string",
                        "description": "Path untuk menentukan manual book yang akan diunduh.",
                        "example": "1",
                    }
                ],
                "responses": {
                    "200": {"description": "Manual book berhasil diunduh."},
                    "500": {"description": "Terjadi kesalahan internal server."},
                },
            }
        },
        "/sign-in": {
            "get": {
                "tags": ["Autentikasi"],
                "summary": "Halaman Masuk Pengguna",
                "description": "Menjalankan halaman autentikasi signin",
                "parameters": [
                    {
                        "name": "msg",
                        "in": "query",
                        "type": "string",
                        "required": False,
                        "description": "Pesan untuk pengguna",
                        "example": "Selamat datang di AbsensiKu",
                    },
                    {
                        "name": "title",
                        "in": "query",
                        "type": "string",
                        "required": False,
                        "description": "Judul Pesan untuk pengguna",
                        "example": "SignIn",
                    },
                    {
                        "name": "status",
                        "in": "query",
                        "type": "string",
                        "enum": ["success"],
                        "required": False,
                        "description": "status pesan untuk pengguna",
                        "example": "Selamat datang di AbsensiKu",
                    },
                ],
                "responses": {
                    "200": {"description": "Halaman Berhasil dimuat"},
                    "404": {"description": "Not Found"},
                },
            },
            "post": {
                "tags": ["Autentikasi"],
                "summary": "Proses Masuk Pengguna",
                "description": "Mengizinkan pengguna untuk masuk ke aplikasi dengan memberikan kredensial yang valid.",
                "consumes": ["application/json"],
                "parameters": [
                    {
                        "in": "body",
                        "name": "body",
                        "description": "Data login pengguna berupa email, password, dan jobs ('Karyawan', 'Magang', 'Admin','Sub Admin').",
                        "required": True,
                        "schema": {
                            "type": "object",
                            "properties": {
                                "email": {
                                    "type": "string",
                                    "example": "user@example.com",
                                },
                                "password": {
                                    "type": "string",
                                    "example": "password123",
                                },
                                "jobs": {
                                    "type": "string",
                                    "enum": [
                                        "Karyawan",
                                        "Magang",
                                        "Admin",
                                        "Sub Admin",
                                    ],
                                    "example": "Karyawan",
                                },
                            },
                        },
                    },
                    {
                        "name": "X-CSRF-Token",
                        "in": "header",
                        "required": True,
                        "schema": {"type": "string", "example": "abc123xyz"},
                    },
                ],
                "responses": {
                    "200": {
                        "description": "Login berhasil.",
                        "schema": {
                            "type": "object",
                            "properties": {
                                "message": {
                                    "type": "string",
                                    "example": "login berhasil",
                                },
                                "status": {"type": "string", "example": "success"},
                            },
                        },
                    },
                    "302": {
                        "description": "Redirect ke halaman dashboard dengan pesan sukses.",
                        "headers": {
                            "Location": {
                                "description": "URL tujuan redirect dengan parameter query.",
                                "type": "string",
                                "example": "/dashboard?msg=login success&status=success&title=SignUp%21",
                            },
                        },
                    },
                    "400": {
                        "description": "Input tidak valid.",
                        "properties": {
                            "type": "object",
                            "message": {
                                "type": "string",
                                "example": "Input tidak valid.",
                            },
                        },
                    },
                    "500": {
                        "description": "Terjadi kesalahan internal server.",
                        "properties": {
                            "type": "object",
                            "message": {
                                "type": "string",
                                "example": "Input tidak valid.",
                            },
                        },
                    },
                },
            },
        },
        "/sign-in/forget": {
            "post": {
                "tags": ["Autentikasi"],
                "summary": "Lupa Password Post",
                "description": "Mengizinkan pengguna untuk mengatur ulang kata sandi mereka. Memerlukan email dan token CSRF yang valid.",
                "consumes": ["application/x-www-form-urlencoded"],
                "parameters": [
                    {
                        "name": "csrf_token",
                        "in": "formData",
                        "required": True,
                        "description": "Token CSRF untuk keamanan.",
                        "type": "string",
                    },
                    {
                        "name": "email",
                        "in": "formData",
                        "required": True,
                        "description": "Alamat email pengguna.",
                        "schema": {"type": "string"},
                    },
                    {
                        "name": "password_new",
                        "in": "formData",
                        "required": True,
                        "description": "Kata sandi baru pengguna.",
                        "schema": {"type": "string"},
                    },
                    {
                        "name": "password2_new",
                        "in": "formData",
                        "required": True,
                        "description": "Konfirmasi kata sandi baru.",
                        "schema": {"type": "string"},
                    },
                ],
                "responses": {
                    "302": {
                        "description": "OTP telah dikirimkan",
                        "schema": {
                            "type": "object",
                            "properties": {
                                "message": {
                                    "type": "string",
                                    "example": "OTP berhasil dikirimkan",
                                },
                                "status": {"type": "string", "example": "success"},
                            },
                        },
                    },
                    "400": {
                        "description": "Input tidak valid.",
                        "schema": {
                            "type": "object",
                            "properties": {
                                "message": {
                                    "type": "string",
                                    "example": "Input tidak valid",
                                },
                                "status": {"type": "string", "example": "error"},
                            },
                        },
                    },
                    "500": {
                        "description": "Terjadi kesalahan internal server.",
                        "schema": {
                            "type": "object",
                            "properties": {
                                "message": {
                                    "type": "string",
                                    "example": "Terjadi kesalahan saat mengatur ulang kata sandi.",
                                },
                                "status": {"type": "string", "example": "error"},
                            },
                        },
                    },
                },
            },
            "get": {
                "tags": ["Autentikasi"],
                "summary": "Halaman lupa password",
                "description": "Menjalankan halaman lupa password",
                "parameters": [
                    {
                        "name": "msg",
                        "in": "query",
                        "type": "string",
                        "required": False,
                        "description": "Pesan untuk pengguna",
                        "example": "OTP Kadaluarsa",
                    },
                ],
                "responses": {
                    "200": {"description": "Halaman berhasil dimuat"},
                    "404": {"description": "Not Found"},
                },
            },
        },
        "/sign-in/forget/otp/<jwt_otp>": {
            "post": {
                "tags": ["Autentikasi"],
                "summary": "Verifikasi OTP untuk Reset Password",
                "description": "Mengizinkan pengguna untuk memverifikasi OTP yang dikirim ke email mereka untuk mengatur ulang kata sandi.",
                "parameters": [
                    {
                        "name": "jwt_otp",
                        "in": "path",
                        "required": True,
                        "description": "Token JWT yang berisi OTP.",
                        "schema": {"type": "string"},
                    },
                    {
                        "in": "body",
                        "name": "body",
                        "required": True,
                        "description": "OTP yang dimasukkan oleh pengguna.",
                        "schema": {
                            "type": "string",
                            "properties": {
                                "otp": {"type": "string", "example": "123456"},
                            },
                        },
                    },
                    {
                        "name": "csrf_token",
                        "in": "header",
                        "required": True,
                        "description": "Token CSRF untuk keamanan.",
                        "type": "string",
                    },
                ],
                "responses": {
                    "200": {
                        "description": "OTP berhasil diverifikasi dan kata sandi diperbarui.",
                        "schema": {
                            "type": "object",
                            "properties": {
                                "redirect": {"type": "string", "example": "/sign-in"},
                                "msg": {
                                    "type": "string",
                                    "example": "Password updated",
                                },
                                "status": {"type": "string", "example": "success"},
                            },
                        },
                    },
                    "400": {
                        "description": "OTP tidak valid.",
                        "schema": {
                            "type": "object",
                            "properties": {
                                "message": {
                                    "type": "string",
                                    "example": "OTP tidak valid",
                                },
                                "status": {"type": "string", "example": "error"},
                            },
                        },
                    },
                    "500": {
                        "description": "Terjadi kesalahan internal server.",
                        "schema": {
                            "type": "object",
                            "properties": {
                                "message": {
                                    "type": "string",
                                    "example": "Terjadi kesalahan saat memverifikasi OTP.",
                                },
                                "status": {"type": "string", "example": "error"},
                            },
                        },
                    },
                },
            },
            "get": {
                "tags": ["Autentikasi"],
                "summary": "Halaman lupa password dengan OTP",
                "description": "Menjalankan halaman untuk memasukkan OTP.",
                "parameters": [
                    {
                        "name": "jwt_otp",
                        "in": "path",
                        "required": True,
                        "description": "Token JWT yang berisi OTP.",
                        "schema": {"type": "string"},
                    }
                ],
                "responses": {
                    "200": {"description": "Halaman berhasil dimuat"},
                    "404": {"description": "Not Found"},
                },
            },
        },
        "/sign-up": {
            "get": {
                "tags": ["Autentikasi"],
                "summary": "Halaman Pendaftaran",
                "description": "Menampilkan halaman pendaftaran untuk pengguna baru.",
                "responses": {
                    "200": {"description": "Halaman pendaftaran berhasil dimuat."}
                },
            },
            "post": {
                "tags": ["Autentikasi"],
                "summary": "Proses Pendaftaran Pengguna",
                "description": "Endpoint ini digunakan untuk mendaftar pengguna baru ke dalam aplikasi.",
                "consumes": ["application/x-www-form-urlencoded"],
                "parameters": [
                    {
                        "name": "csrf_token",
                        "in": "formData",
                        "required": True,
                        "description": "Token CSRF untuk keamanan.",
                        "type": "string",
                        "example": "IjFhMGU1ODI0NDA0Nzk1MjdkOWM4NWM1ZmU4MGJkNWZiNTM0MDdjMjci.Z2lUCQ.jTybc72d55jThpGfqU9ey_p7C1I",
                    },
                    {
                        "name": "nama",
                        "in": "formData",
                        "required": True,
                        "description": "Nama lengkap pengguna.",
                        "schema": {"type": "string"},
                    },
                    {
                        "name": "departement",
                        "in": "formData",
                        "required": True,
                        "enum": [
                            "Fullstack Developer",
                            "Web Developer Laravel",
                            "Copywriting",
                        ],
                        "description": "Departemen pengguna.",
                        "schema": {"type": "string"},
                    },
                    {
                        "name": "email",
                        "in": "formData",
                        "required": True,
                        "description": "Alamat email pengguna.",
                        "schema": {"type": "string"},
                    },
                    {
                        "name": "password",
                        "in": "formData",
                        "required": True,
                        "description": "Kata sandi untuk akun baru.",
                        "schema": {"type": "string"},
                    },
                    {
                        "name": "jobs",
                        "in": "formData",
                        "required": True,
                        "enum": ["Karyawan", "Magang"],
                        "description": "Jabatan pengguna (misalnya, 'Karyawan', 'Magang').",
                        "schema": {"type": "string"},
                    },
                ],
                "responses": {
                    "200": {
                        "description": "Pendaftaran berhasil.",
                        "schema": {
                            "type": "object",
                            "properties": {
                                "message": {
                                    "type": "string",
                                    "example": "Pendaftaran berhasil, silakan masuk.",
                                },
                                "status": {"type": "string", "example": "success"},
                            },
                        },
                    },
                    "302": {
                        "description": "Redirect ke halaman Sign In dengan pesan sukses.",
                        "headers": {
                            "Location": {
                                "description": "URL tujuan redirect dengan parameter query.",
                                "type": "string",
                                "example": "/sign-In?msg=Sign+Up+Success&status=success&title=SignUp%21",
                            },
                        },
                    },
                    "400": {
                        "description": "Permintaan tidak valid atau parameter yang diperlukan hilang.",
                        "schema": {
                            "type": "object",
                            "properties": {
                                "message": {
                                    "type": "string",
                                    "example": "Nama tidak boleh kosong.",
                                },
                            },
                        },
                    },
                    "500": {
                        "description": "Terjadi kesalahan di server.",
                        "schema": {
                            "type": "object",
                            "properties": {
                                "message": {
                                    "type": "string",
                                    "example": "Terjadi kesalahan saat mendaftar.",
                                },
                            },
                        },
                    },
                },
            },
        },
        "/api/auth/logout": {
            "get": {
                "tags": ["Autentikasi"],
                "summary": "Logout Pengguna",
                "description": "Menghapus sesi pengguna dan menghapus cookie autentikasi.",
                "responses": {
                    "200": {
                        "description": "Logout berhasil.",
                        "schema": {
                            "type": "object",
                            "properties": {
                                "status": {"type": "string", "example": "success"},
                                "redirect": {"type": "string", "example": "/sign-in"},
                                "msg": {
                                    "type": "string",
                                    "example": "Logout berhasil.",
                                },
                            },
                        },
                    }
                },
            },
        },
        "/ask": {
            "post": {
                "tags": ["Bantuan"],
                "summary": "Ajukan Pertanyaan",
                "description": "Mengizinkan pengguna untuk mengajukan pertanyaan atau masalah yang mereka hadapi.",
                "parameters": [
                    {
                        "name": "X-CSRF-TOKEN",
                        "in": "header",
                        "required": True,
                        "description": "Token CSRF untuk keamanan.",
                        "type": "string",
                        "example": "IjFhMGU1ODI0NDA0Nzk1MjdkOWM4NWM1ZmU4MGJkNWZiNTM0MDdjMjci.Z2lUCQ.jTybc72d55jThpGfqU9ey_p7C1I",
                    },
                    {
                        "name": "body",
                        "in": "body",
                        "required": True,
                        "description": "Berisi nama, email, jobs, departement, dan kendala.",
                        "schema": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string", "example": "John Doe"},
                                "email": {
                                    "type": "string",
                                    "example": "pZf9M@gmail.com",
                                },
                                "jobs": {
                                    "type": "string",
                                    "example": "Karyawan",
                                    "enum": [
                                        "Karyawan",
                                        "Magang",
                                        "Admin",
                                        "Sub Admin",
                                    ],
                                },
                                "departement": {
                                    "type": "string",
                                    "example": "Fullstack Developer",
                                    "enum": [
                                        "Fullstack Developer",
                                        "Web Developer Laravel",
                                        "Copywriting",
                                    ],
                                },
                                "issue": {"type": "string", "example": "Kendala 1"},
                            },
                        },
                    },
                ],
                "responses": {
                    "200": {
                        "description": "Pertanyaan berhasil diajukan.",
                        "schema": {
                            "type": "object",
                            "properties": {
                                "redirect": {"type": "string"},
                                "msg": {"type": "string"},
                            },
                        },
                    },
                    "400": {
                        "description": "Input tidak valid.",
                        "schema": {
                            "type": "object",
                            "properties": {"message": {"type": "string"}},
                        },
                    },
                    "500": {
                        "description": "Terjadi kesalahan internal server.",
                        "schema": {
                            "type": "object",
                            "properties": {"message": {"type": "string"}},
                        },
                    },
                },
            }
        },
        "/myProfiles": {
            "get": {
                "tags": ["Admin / Sub Admin", "Magang / Karyawan"],
                "summary": "Lihat Profil Pengguna",
                "description": "Endpoint untuk mendapatkan informasi profil pengguna.",
                "parameters": [
                    {
                        "name": "msg",
                        "in": "query",
                        "required": False,
                        "description": "Pesan untuk pengguna.",
                        "type": "string",
                        "example": "Profil berhasil ditampilkan.",
                    },
                    {
                        "name": "status",
                        "in": "query",
                        "required": False,
                        "description": "Status untuk pengguna.",
                        "type": "string",
                        "example": "success",
                    },
                    {
                        "name": "token_key",
                        "in": "cookie",
                        "required": True,
                        "description": "Token untuk keamanan.",
                        "type": "string",
                    },
                    {
                        "name": "csrf_token",
                        "in": "cookie",
                        "required": True,
                        "description": "Token CSRF untuk keamanan.",
                        "type": "string",
                        "example": "IjFhMGU1ODI0NDA0Nzk1MjdkOWM4NWM1ZmU4MGJkNWZiNTM0MDdjMjci.Z2lZcw.GjsOiVWeyKP1DjYSaxXHrPGDmHM",
                    },
                ],
                "responses": {
                    "200": {"description": "Profil pengguna berhasil ditampilkan."},
                    "400": {"description": "Permintaan tidak valid."},
                    "500": {"description": "Terjadi kesalahan internal server."},
                },
            },
            "post": {
                "tags": ["Admin / Sub Admin", "Magang / Karyawan"],
                "summary": "Perbarui Profil Pengguna",
                "description": "Endpoint untuk memperbarui informasi profil pengguna.",
                "parameters": [
                    {
                        "name": "token_key",
                        "in": "cookie",
                        "required": True,
                        "description": "Token untuk keamanan.",
                        "type": "string",
                    },
                    {
                        "name": "csrf_token",
                        "in": "formData",
                        "required": True,
                        "description": "Token CSRF untuk validasi keamanan.",
                        "type": "string",
                        "example": "IjFhMGU1ODI0NDA0Nzk1MjdkOWM4NWM1ZmU4MGJkNWZiNTM0MDdjMjci.Z2lZcw.GjsOiVWeyKP1DjYSaxXHrPGDmHM",
                    },
                    {
                        "name": "email",
                        "in": "formData",
                        "required": True,
                        "description": "Alamat email pengguna.",
                        "type": "string",
                        "example": "admin@gmail.com",
                    },
                    {
                        "name": "nama",
                        "in": "formData",
                        "required": True,
                        "description": "Nama lengkap pengguna.",
                        "type": "string",
                        "example": "john doe",
                    },
                    {
                        "name": "nik",
                        "in": "formData",
                        "required": False,
                        "description": "Nomor Induk Kependudukan (NIK) pengguna. (Hanya untuk role tertentu)",
                        "type": "integer",
                        "example": "12238398131313813",
                    },
                    {
                        "name": "tempat_lahir",
                        "in": "formData",
                        "required": False,
                        "description": "Tempat lahir pengguna. (Hanya untuk role tertentu)",
                        "type": "string",
                        "example": "Jakarta",
                    },
                    {
                        "name": "tanggal_lahir",
                        "in": "formData",
                        "required": False,
                        "description": "Tanggal lahir pengguna dalam format YYYY-MM-DD. (Hanya untuk role tertentu)",
                        "type": "string",
                        "format": "date",
                        "example": "2010-08-12",
                    },
                    {
                        "name": "profile-pic",
                        "in": "formData",
                        "required": False,
                        "description": "Gambar profil pengguna.",
                        "type": "file",
                        "example": "datas.png",
                    },
                ],
                "responses": {
                    "200": {"description": "Profil berhasil diperbarui."},
                    "400": {"description": "Input tidak valid."},
                    "500": {"description": "Terjadi kesalahan internal server."},
                },
            },
        },
        "/change-password": {
            "get": {
                "tags": ["Admin / Sub Admin", "Magang / Karyawan"],
                "summary": "Halaman Ubah Password Pengguna",
                "description": "Mengizinkan pengguna untuk mengakses halaman ubah password mereka.",
                "parameters": [
                    {
                        "name": "token_key",
                        "in": "cookie",
                        "type": "string",
                        "required": True,
                        "description": "token key dari cookie",
                    },
                    {
                        "name": "csrf_token",
                        "in": "cookie",
                        "type": "string",
                        "required": True,
                        "description": "csrf dari cookie",
                    },
                    {
                        "name": "result",
                        "in": "query",
                        "type": "string",
                        "enum": ["success"],
                        "required": False,
                        "description": "result dari untuk pesan",
                        "example": "success",
                    },
                    {
                        "name": "msg",
                        "in": "query",
                        "type": "string",
                        "required": False,
                        "description": "result dari untuk pesan",
                        "example": "ini merupakan pesan dari error",
                    },
                ],
                "responses": {
                    "200": {
                        "description": "Halaman ubah password berhasil ditampilkan."
                    },
                    "401": {"description": "Akses tidak sah atau sesi kedaluwarsa."},
                },
            },
            "post": {
                "tags": ["Admin / Sub Admin", "Magang / Karyawan"],
                "summary": "Ubah Password Pengguna",
                "description": "Mengizinkan pengguna untuk mengubah kata sandi lama mereka ke yang baru.",
                "consumes": ["application/x-www-form-urlencoded"],
                "parameters": [
                    {
                        "name": "token_key",
                        "in": "cookie",
                        "type": "string",
                        "required": True,
                        "description": "token key dari cookie",
                    },
                    {
                        "name": "csrf_token",
                        "in": "cookie",
                        "type": "string",
                        "required": True,
                        "description": "csrf dari cookie",
                    },
                    {
                        "name": "X-CSRFToken",
                        "in": "header",
                        "type": "string",
                        "required": True,
                        "description": "csrf dari cookie",
                    },
                    {
                        "name": "old_password",
                        "in": "formData",
                        "type": "string",
                        "required": True,
                        "description": "Kata sandi saat ini pengguna.",
                    },
                    {
                        "name": "new_password",
                        "in": "formData",
                        "type": "string",
                        "required": True,
                        "description": "Kata sandi baru yang diinginkan pengguna.",
                    },
                ],
                "responses": {
                    "200": {"description": "Password berhasil diubah."},
                    "400": {
                        "description": "Input tidak valid atau kata sandi tidak cocok."
                    },
                    "401": {
                        "description": "Kesalahan autentikasi (sesi/token kedaluwarsa)."
                    },
                },
            },
        },
        "/riwayat-kehadiran/{path1}": {
            "post": {
                "tags": ["Admin / Sub Admin"],
                "summary": "Mengelola Riwayat Kehadiran",
                "description": "Mengelola riwayat kehadiran pengguna berdasarkan parameter yang diberikan yaitu edit.",
                "parameters": [
                    {
                        "name": "path1",
                        "in": "path",
                        "required": True,
                        "type": "string",
                        "enum": ["edit", "delete"],
                        "description": "Tindakan yang akan dilakukan (edit/delete).",
                    },
                    {
                        "name": "csrf_token",
                        "in": "cookie",
                        "required": True,
                        "description": "Token CSRF yang valid.",
                        "type": "string",
                    },
                    {
                        "name": "token_key",
                        "in": "cookie",
                        "required": True,
                        "description": "Token Key yang valid.",
                        "type": "string",
                    },
                    {
                        "name": "__method",
                        "in": "formData",
                        "required": True,
                        "description": "Input method [PUT / DELETE]",
                        "enum": ["PUT", "DELETE"],
                        "type": "string",
                    },
                    {
                        "name": "__csrf_token",
                        "in": "formData",
                        "required": True,
                        "description": "Input csrf form roken",
                        "type": "string",
                    },
                    {
                        "name": "__id_riwayat_absent",
                        "in": "formData",
                        "required": False,
                        "description": "Input id riwayat absen (required for edit)",
                        "type": "string",
                    },
                    {
                        "name": "nik",
                        "in": "formData",
                        "required": False,
                        "description": "Input nik user ( required for edit)",
                        "type": "string",
                    },
                    {
                        "name": "email",
                        "in": "formData",
                        "required": False,
                        "description": "Input email user ( required for edit)",
                        "type": "string",
                    },
                    {
                        "name": "status_hadir",
                        "in": "formData",
                        "required": False,
                        "enum": ["1", "2", "0", "3"],
                        "description": "Input status_hadir user ( required for edit)",
                        "type": "string",
                        "example": "1",
                    },
                ],
                "responses": {
                    "200": {"description": "Riwayat kehadiran berhasil dikelola."},
                    "400": {"description": "Input tidak valid."},
                    "401": {
                        "description": "Token CSRF atau cookie tidak valid atau telah kedaluwarsa."
                    },
                    "500": {"description": "Terjadi kesalahan di server."},
                },
            }
        },
        "/riwayat-kehadiran/{path1}/{path2}": {
            "post": {
                "tags": ["Admin / Sub Admin"],
                "summary": "Mengelola Riwayat Kehadiran",
                "description": "Mengelola riwayat kehadiran pengguna berdasarkan parameter yang diberikan yaitu delete.",
                "parameters": [
                    {
                        "name": "path1",
                        "in": "path",
                        "required": True,
                        "type": "string",
                        "enum": ["edit", "delete"],
                        "description": "Tindakan yang akan dilakukan (edit/delete).",
                    },
                    {
                        "name": "path2",
                        "in": "path",
                        "required": True,
                        "type": "string",
                        "description": "input uuid",
                    },
                    {
                        "name": "csrf_token",
                        "in": "cookie",
                        "required": True,
                        "description": "Token CSRF yang valid.",
                        "type": "string",
                    },
                    {
                        "name": "token_key",
                        "in": "cookie",
                        "required": True,
                        "description": "Token Key yang valid.",
                        "type": "string",
                    },
                    {
                        "name": "__method",
                        "in": "formData",
                        "required": True,
                        "description": "Input method [PUT / DELETE]",
                        "enum": ["PUT", "DELETE"],
                        "type": "string",
                    },
                    {
                        "name": "__csrf_token",
                        "in": "formData",
                        "required": True,
                        "description": "Input csrf form roken",
                        "type": "string",
                    },
                ],
                "responses": {
                    "200": {"description": "Riwayat kehadiran berhasil dikelola."},
                    "400": {"description": "Input tidak valid."},
                    "401": {
                        "description": "Token CSRF atau cookie tidak valid atau telah kedaluwarsa."
                    },
                    "500": {"description": "Terjadi kesalahan di server."},
                },
            }
        },
        "/riwayat-bantuan": {
            "get": {
                "tags": ["Admin / Sub Admin"],
                "summary": "Riwayat Bantuan",
                "description": "Menampilkan riwayat bantuan yang diajukan oleh pengguna.",
                "parameters": [
                    {
                        "name": "token_key",
                        "in": "cookie",
                        "type": "string",
                        "required": True,
                        "description": "token key dari cookie",
                    },
                    {
                        "name": "csrf_token",
                        "in": "cookie",
                        "type": "string",
                        "required": True,
                        "description": "csrf dari cookie",
                    },
                    {
                        "name": "status",
                        "in": "query",
                        "type": "string",
                        "enum": ["success"],
                        "required": False,
                        "description": "result dari untuk pesan",
                        "example": "success",
                    },
                    {
                        "name": "msg",
                        "in": "query",
                        "type": "string",
                        "required": False,
                        "description": "result dari untuk pesan",
                        "example": "ini merupakan pesan dari error",
                    },
                ],
                "responses": {
                    "200": {"description": "Berhasil menampilkan riwayat bantuan."},
                    "401": {"description": "Akses ditolak."},
                    "500": {"description": "Terjadi kesalahan server."},
                },
            }
        },
        "/update-status-bantuan": {
            "post": {
                "tags": ["Admin / Sub Admin"],
                "summary": "Perbarui Status Bantuan",
                "description": "Endpoint ini digunakan untuk memperbarui status permintaan bantuan yang diajukan oleh pengguna.",
                "parameters": [
                    {
                        "name": "body",
                        "in": "body",
                        "required": True,
                        "description": "Body permintaan yang berisi status dan ID permintaan bantuan yang akan diperbarui.",
                        "schema": {
                            "type": "object",
                            "properties": {
                                "status": {
                                    "type": "string",
                                    "description": "Status baru dari permintaan bantuan.",
                                    "enum": ["Selesai", "Pending", "Diproses"],
                                    "default": "Pending",
                                    "example": "Selesai",
                                },
                                "status_id": {
                                    "type": "string",
                                    "description": "ID permintaan bantuan yang akan diperbarui.",
                                    "example": "5f4e2b2a2f8d4e38b2d0f5b1",
                                },
                            },
                        },
                    },
                    {
                        "name": "token_key",
                        "in": "cookie",
                        "type": "string",
                        "required": True,
                        "description": "token key dari cookie",
                    },
                    {
                        "name": "csrf_token",
                        "in": "cookie",
                        "type": "string",
                        "required": True,
                        "description": "csrf dari cookie",
                    },
                    {
                        "name": "X-Requested-With",
                        "in": "header",
                        "type": "string",
                        "required": True,
                        "description": "requested with",
                        "example": "XMLHttpRequest",
                    },
                    {
                        "name": "X-CSRF-TOKEN",
                        "in": "header",
                        "type": "string",
                        "required": True,
                        "description": "input X-CSRF-TOKEN",
                        "example": "your-csrf-token",
                    },
                ],
                "responses": {
                    "200": {
                        "description": "Status permintaan bantuan berhasil diperbarui."
                    },
                    "400": {"description": "Permintaan tidak valid."},
                    "401": {
                        "description": "Token CSRF atau cookie tidak valid atau telah kedaluwarsa."
                    },
                    "500": {"description": "Terjadi kesalahan di server."},
                },
            }
        },
        "/kelola-admin": {
            "get": {
                "tags": ["Admin / Sub Admin"],
                "summary": "Halaman Kelola Admin",
                "description": "Menampilkan halaman untuk mengelola admin dan sub-admin.",
                "parameters": [
                    {
                        "name": "token_key",
                        "in": "cookie",
                        "type": "string",
                        "required": True,
                        "description": "token key dari cookie",
                    },
                    {
                        "name": "csrf_token",
                        "in": "cookie",
                        "type": "string",
                        "required": True,
                        "description": "csrf dari cookie",
                    },
                    {
                        "name": "status",
                        "in": "query",
                        "type": "string",
                        "enum": ["success"],
                        "required": False,
                        "description": "result dari untuk pesan",
                        "example": "success",
                    },
                    {
                        "name": "msg",
                        "in": "query",
                        "type": "string",
                        "required": False,
                        "description": "result dari untuk pesan",
                        "example": "ini merupakan pesan dari error",
                    },
                ],
                "responses": {
                    "200": {"description": "Halaman berhasil dimuat."},
                    "401": {"description": "Akses tidak sah."},
                    "500": {"description": "Terjadi kesalahan internal server."},
                },
            }
        },
        "/kelola-admin/": {
            "get": {
                "tags": ["Admin / Sub Admin"],
                "summary": "Halaman Kelola Admin (dengan trailing slash)",
                "description": "Menampilkan halaman untuk mengelola admin dan sub-admin.",
                "parameters": [
                    {
                        "name": "token_key",
                        "in": "cookie",
                        "type": "string",
                        "required": True,
                        "description": "token key dari cookie",
                    },
                    {
                        "name": "csrf_token",
                        "in": "cookie",
                        "type": "string",
                        "required": True,
                        "description": "csrf dari cookie",
                    },
                    {
                        "name": "result",
                        "in": "query",
                        "type": "string",
                        "enum": ["success"],
                        "required": False,
                        "description": "result dari untuk pesan",
                        "example": "success",
                    },
                    {
                        "name": "msg",
                        "in": "query",
                        "type": "string",
                        "required": False,
                        "description": "result dari untuk pesan",
                        "example": "ini merupakan pesan dari error",
                    },
                ],
                "responses": {
                    "200": {"description": "Halaman berhasil dimuat."},
                    "401": {"description": "Akses tidak sah."},
                    "500": {"description": "Terjadi kesalahan internal server."},
                },
            }
        },
        "/kelola-admin/<path1>": {
            "post": {
                "tags": ["Admin / Sub Admin"],
                "summary": "Edit Admin / Sub Admin",
                "description": "Endpoint ini digunakan untuk mengedit  data admin atau sub-admin.",
                "parameters": [
                    {
                        "name": "token_key",
                        "in": "cookie",
                        "type": "string",
                        "required": True,
                        "description": "token key dari cookie",
                    },
                    {
                        "name": "csrf_token",
                        "in": "cookie",
                        "type": "string",
                        "required": True,
                        "description": "csrf dari cookie",
                    },
                    {
                        "name": "path1",
                        "in": "path",
                        "required": True,
                        "description": "Operasi yang akan dilakukan, bisa 'edit' atau 'delete'.",
                        "enum": ["edit", "delete"],
                        "example": "edit",
                        "type": "string",
                    },
                    {
                        "name": "__csrf_token",
                        "in": "formData",
                        "required": True,
                        "description": "Token CSRF Form yang valid.",
                        "type": "string",
                    },
                    {
                        "name": "__method",
                        "in": "formData",
                        "required": True,
                        "description": "method Form yang valid.",
                        "enum": ["PUT", "DELETE"],
                        "type": "string",
                    },
                    {
                        "name": "__id_data_user_admin",
                        "in": "formData",
                        "required": True,
                        "description": "method id user yang valid.",
                        "type": "string",
                    },
                    {
                        "name": "nama",
                        "in": "formData",
                        "required": True,
                        "description": "Nama admin yang akan diedit (hanya untuk edit).",
                        "type": "string",
                    },
                    {
                        "name": "email",
                        "in": "formData",
                        "required": True,
                        "description": "Alamat email admin yang akan diedit (hanya untuk edit).",
                        "type": "string",
                    },
                    {
                        "name": "jobs",
                        "in": "formData",
                        "required": True,
                        "description": "Jabatan admin yang akan diedit (hanya untuk edit).",
                        "type": "string",
                    },
                    {
                        "name": "departement",
                        "in": "formData",
                        "required": True,
                        "description": "Departemen admin yang akan diedit (hanya untuk edit).",
                        "type": "string",
                    },
                ],
                "responses": {
                    "200": {
                        "description": "Data admin berhasil diperbarui atau dihapus."
                    },
                    "400": {"description": "Permintaan tidak valid."},
                    "401": {"description": "Token CSRF atau cookie tidak valid."},
                    "500": {"description": "Terjadi kesalahan di server."},
                },
            }
        },
        "/kelola-admin/<path1>/<path2>": {
            "post": {
                "tags": ["Admin / Sub Admin"],
                "summary": "Operasi delete pada Admin / Sub Admin",
                "description": "Endpoint ini digunakan untuk melakukan operasi delete pada admin atau sub-admin berdasarkan parameter yang diberikan.",
                "parameters": [
                    {
                        "name": "path1",
                        "in": "path",
                        "required": True,
                        "type": "string",
                        "enum": ["edit", "delete"],
                        "description": "Tindakan yang akan dilakukan (edit/delete).",
                    },
                    {
                        "name": "path2",
                        "in": "path",
                        "required": True,
                        "type": "string",
                        "description": "input uuid",
                    },
                    {
                        "name": "csrf_token",
                        "in": "cookie",
                        "required": True,
                        "description": "Token CSRF yang valid.",
                        "type": "string",
                    },
                    {
                        "name": "token_key",
                        "in": "cookie",
                        "required": True,
                        "description": "Token Key yang valid.",
                        "type": "string",
                    },
                    {
                        "name": "__method",
                        "in": "formData",
                        "required": True,
                        "description": "Input method [PUT / DELETE]",
                        "enum": ["PUT", "DELETE"],
                        "type": "string",
                    },
                    {
                        "name": "__csrf_token",
                        "in": "formData",
                        "required": True,
                        "description": "Input csrf form roken",
                        "type": "string",
                    },
                ],
                "responses": {
                    "200": {"description": "Operasi berhasil dilakukan."},
                    "400": {"description": "Permintaan tidak valid."},
                    "401": {"description": "Token CSRF atau cookie tidak valid."},
                    "500": {"description": "Terjadi kesalahan di server."},
                },
            }
        },
    },
}

swagger = Swagger(app, template=swagger_template)


# default url dokumentasi
@app.before_request
def update_host_for_api_docs():
    # Cek jika URL yang diakses adalah /api_docs/
    if request.path == "/apidocs/":
        global swagger_template
        host_name = request.host
        # Update template global
        swagger_template["host"] = host_name
        # Update Swagger instance
        app.config["SWAGGER"] = {"template": swagger_template}


# home
@app.route("/", methods=["GET"])
def home():
    """
    Home Page
    -

    tags:
      - Landing Page

    This function will return the home.html which is the home page of the application.
    The home page will redirect to the signIn page if the user is not authenticated.
    """
    return render_template("home.html")


# manual book
@app.route("/manual/<path>", methods=["GET"])
def manual_book(path):
    """
    Download Manual Book
    -

    This function will return the manual book of the application.
    The manual book is a PDF file that contains the user manual of the application.
    The function will check if the path is valid and if the file exists.
    If the file exists, it will return the file as a PDF attachment.
    If the file does not exist, it will return a 500 error.

    tags:
        - Export

    Parameters:
    -
    path : str
        The path of the manual book to download.
        The path should be a string that contains the version number of the manual book.
        The version number should be a single digit number.

    Returns:
    -
    A PDF file that contains the user manual of the application.
    """
    file_path = os.path.join(app.root_path, "static", "doc")
    if not file_path:
        return make_response(jsonify({"redirect": url_for("home")}), 500)
    if path == "1":
        return send_file(
            file_path + "/pdf/manual_admin_subadmin.pdf",
            mimetype="application/pdf",
            as_attachment=True,
            download_name="manual_admin_subadmin_winnicode.pdf",
        )
    elif path == "3":
        return send_file(
            file_path + "/pdf/manual_karyawan_magang.pdf",
            mimetype="application/pdf",
            as_attachment=True,
            download_name="manual_karyawan_magang_winnicode.pdf",
        )
    else:
        return make_response(jsonify({"redirect": url_for("home")}), 500)


# lakukan sign-in
@app.route("/sign-in", methods=["GET", "POST"])
@app.route("/sign-in/", methods=["GET", "POST"])
def signIn():
    """
    Sign In Function
    -

    This function handles the user sign-in process for the application. It supports both GET and POST
    methods to facilitate the sign-in workflow.

    tags:
        - Autentikasi

    Usage:
    -
    1. Send a POST request to the "/sign-in" endpoint with the following JSON payload:
        {
            "email": "<user_email>",
            "password": "<user_password>",
            "jobs": "<user_jobs>"
        }
        Additionally, include the CSRF token in the request headers as "X-CSRF-Token".

    2. If the email field is empty, an exception will be raised indicating that the email cannot be
        empty.

    Parameters:
    -
    - email: str
        The user's email address used for authentication.
    - password: str
        The user's password for authentication.
    - jobs: str
        The user's job or position, required for role-based access control.
    - csrf_token: str
        The CSRF token for security validation, passed in the request headers.

    Returns:
    -
    A response indicating the success or failure of the sign-in attempt. On success, the user is
    authenticated and redirected to their dashboard. On failure, an appropriate error message is
    returned.

    Note:
    -
    Ensure that the CSRF token is valid and included in the request headers to avoid security errors.

    """

    if request.method == "POST":
        try:
            email, password, jobs, csrf_token = (
                request.json.get("email"),
                request.json.get("password"),
                request.json.get("jobs"),
                request.headers.get("X-CSRF-Token"),
            )
            if email == "":
                raise Exception("Email tidak boleh kosong")
            elif password == "":
                raise Exception("Password tidak boleh kosong")
            elif csrf_token == "":
                raise Exception("CSRF Token tidak boleh kosong")
            elif jobs == "None":
                raise Exception("Jobs tidak boleh kosong")
            else:
                user = db["users"].find_one(
                    {
                        "email": email.lower(),
                        "jobs": jobs,
                    }
                )
                csrf_token = string_to_uuid_like(csrf_token)

                if not csrf_token:
                    raise Exception("encoding Gagal")

                # cek user
                if not user:
                    raise Exception("Email Salah")

                # cek password sama / tidak
                if not (
                    user["password"] == hashlib.sha256(password.encode()).hexdigest()
                ):
                    raise Exception("Password Salah")

                # cek jobs
                if user["jobs"] in ("Magang", "Karyawan"):
                    # cek mulai kerja dan akhir kerja ada/ tidak ada
                    if user["mulai_kerja"] and user["akhir_kerja"]:
                        # jika tanggal kerja tidak sesuai
                        if not cek_tanggal_kerja(
                            user["mulai_kerja"], user["akhir_kerja"]
                        ):
                            raise Exception(
                                "Akun kamu kadaluarsa atau waktu bekerjamu belum dimulai!"
                            )
                    else:
                        raise Exception("Akun mu pending, tunggu sampai 24 jam")
                elif user["jobs"] in ("Admin", "Sub Admin"):
                    pass
                else:
                    raise Exception("Salah jobs gan!!")

                # make token jwt
                token = signInPayload(
                    user["_id"],
                    user["jobs"],
                    user["role"],
                    datetime.timedelta(minutes=30),
                )

                # make the jwt
                token = string_to_uuid_like(token)

                # membuat response jsonify true
                resp = make_response(
                    jsonify(
                        {
                            "result": "success",
                            "redirect": "/dashboard",
                            "msg": "Login berhasil!",
                        }
                    ),
                    200,
                )
                
                print(get_time_zone_now())

                # set cookie
                resp.set_cookie(
                    "token_key",
                    token,
                    httponly=True,
                    secure=True,
                    samesite="Lax",
                    expires=get_time_zone_now() + datetime.timedelta(minutes=30),
                )  # ubah secure jadi true saat production
                resp.set_cookie(
                    "csrf_token",
                    csrf_token,
                    httponly=True,
                    secure=True,
                    samesite="Lax",
                    expires=get_time_zone_now() + datetime.timedelta(minutes=30),
                )  # ubah secure jadi true saat production

                # kembalikan responsenya
                return resp

        # handle error
        except Exception as e:
            return jsonify({"redirect": url_for("signIn", msg=e.args[0])}), 500

    # request method get
    msg = request.args.get("msg")
    status = request.args.get("status")
    title = request.args.get("title")

    # ambil dat aadmin keseluruhan
    admin_user_cek = list(
        db.users.find({"jobs": "Admin", "departement": "Superuser", "role": 1})
    )
    # cek apakah memiliki user admin / tidak?
    # jika tidak buat default awalan untuk superuser
    if not admin_user_cek:
        result = db.users.insert_one(
            {
                "nama": "admin",
                "email": "admin@gmail.com".lower(),
                "password": hashlib.sha256("admin123".strip().encode()).hexdigest(),
                "departement": "Superuser".strip(),
                "jobs": "Admin".strip(),
                "role": 1,
                "photo_profile": "https://i.ibb.co.com/5Yd94zx/user.png",
            }
        )
        # jika error ditengah tengah
        if not result:
            msg = "Gagal membuat default(admin)"
    return render_template("signIn.html", msg=msg, status=status, title=title)


# lakukan Add Karyawan dan Magang
@app.route("/dashboard/admin/add", methods=["POST"])
@app.route("/dashboard/admin/add/", methods=["POST"])
def karyawanMagangAdd():
    """
    fungsi ini digunakan untuk mendaftarkan Karyawan / Magang.

    parameters :
        - csrf_token : token csrf
        - nama : nama Karyawan / Magang
        - departement : departemen Karyawan / Magang
        - email : email Karyawan / Magang
        - password : password Karyawan / Magang
        - jobs : posisi Karyawan / Magang

    responses :
        - 200 : berhasil mendaftarkan Karyawan / Magang
        - 400 : inputan tidak valid
        - 401 : csrf token tidak valid
        - 500 : terjadi kesalahan di server
    """
    try:
        if request.method == "POST":
            # cek cookie token
            cookies = request.cookies.get("token_key")
            
            # decode uuid
            cookies = uuid_like_to_string(cookies)

            # cek berhasul / tidak
            if not cookies:
                raise ValueError("CSRF decode error")
            # decode jwt
            payloads = jwt.decode(cookies, secretKey, algorithms=["HS256"])
            
            data_user = db.users.find_one(
                {"_id": ObjectId(payloads["_id"])},
                {"_id": 0, "departement": 1, "jobs": 1, "role": 1},
            )

            if not data_user:
                raise Exception("Data user akses tidak ditemukan")
            
            # cek hak akses selama cookie
            if (
                payloads["role"] != 1
                and payloads["jobs"] not in ("Admin",'Sub Admin')
                and data_user["departement"] not in ("Superuser")
            ):
                raise Exception("Anda tidak memiliki akses")

            csrf_token, nama, nik, email, departement, jobs, password, password2  = request.form.values()
            nik = int(nik)
            if nama == "":
                return redirect(url_for("dashboard", msg="Nama tidak boleh kosong"))
            if nik == "":
                return redirect(url_for("dashboard", msg="Nik tidak boleh kosong"))
            elif departement == "None":
                return redirect(url_for("dashboard", msg="Departement tidak boleh kosong"))
            elif jobs == "None":
                return redirect(url_for("dashboard", msg="Jobs tidak boleh kosong"))
            elif email == "":
                return redirect(url_for("dashboard", msg="Email tidak boleh kosong"))
            elif password == "":
                return redirect(url_for("dashboard", msg="Password tidak boleh kosong"))
            elif csrf_token == "":
                return redirect(url_for("dashboard", msg="csrf token tidak boleh kosong"))

            if not (password == password2):
                return redirect(url_for("dashboard", msg="Password tidak sama"))
            if not type(nik) == int:
                return redirect(url_for("dashboard", msg="NIK wajib numeric"))
        
            cek_data = db.users.find_one({"email": email.lower()})
            if cek_data:
                return redirect(url_for("dashboard", msg="Akun sudah ada"))
            result = db.users.insert_one(
                {
                    "nama": nama,
                    "departement": departement,
                    "jobs": jobs,
                    "email": email.lower(),
                    "password": hashlib.sha256(password.encode()).hexdigest(),
                    "role": 3,
                    "nik": int(nik),
                    "photo_profile": "https://i.ibb.co.com/5Yd94zx/user.png",
                    "tempat_lahir": "",
                    "tanggal_lahir": "",
                    "mulai_kerja": "",
                    "akhir_kerja": "",
                    "waktu_awal_kerja": "",
                    "waktu_akhir_kerja": "",
                    "work_hours": 0,
                    "absen": {"hadir": 0, "telat": 0, "tidak_hadir": 0, "libur": 0},
                }
            )
            if result:
                return redirect(
                    url_for(
                        "dashboard",
                        msg="Success bertambah akun",
                        result="success",
                    )
                )
            else:
                return Exception('Data {} gagal ditambah'.format(nama))
    # handle error
    except jwt.ExpiredSignatureError:
        return redirect(url_for("signIn", msg="Session Kadaluarsa"))
    except jwt.DecodeError:
        return redirect(url_for("signIn", msg="Anda telah logout"))
    except ValueError as e:
        return redirect(url_for("signIn", msg=e.args[0]))
    except Exception as e:
         return redirect(url_for("dashboard", msg=e.args[0]))


# logout
@app.route("/api/auth/logout", methods=["GET"])
def signOut():
    """
    Logout API
    -

    This function is used to logout from the application.

    Parameters:
    -
        - None

    Returns:
    -
        - JSON response with status, redirect and message
        - status (string): Status of the response
        - redirect (string): URL to redirect after logout
        - msg (string): Message after logout

    Example:
    -
        - GET /api/auth/logout

    """
    # Function implementation continues...
    resp = make_response(
        jsonify(
            {"status": "success", "redirect": "/sign-in", "msg": "Logout berhasil!!"}
        ),
        200,
    )
    # lakukan hapus cookie
    resp.set_cookie("token_key", expires=0)
    resp.set_cookie("csrf_token", expires=0)
    resp.set_cookie("session", expires=0)
    return resp


# lupa password get and post
@app.route("/sign-in/forget", methods=["GET", "POST"])
def forgetPassword():
    """
    Lupa Password
    -
    tags:
        - Autentikasi

    Endpoint ini digunakan untuk mengubah password lama dengan password yang baru. Memerlukan input OTP yang dikirimkan melalui email.

    Parameters:
        - csrf_token (string): CSRF token
        - email (string): Email pengguna
        - password_new (string): Password yang ingin diubah
        - password2_new (string): Konfirmasi password yang ingin diubah

    Returns:
        - Redirect ke halaman login jika berhasil
        - Redirect ke halaman lupa password dengan pesan error jika gagal
    """
    if request.method == "POST":
        try:
            # inisisasi data
            csrf_token, email, password_new, password2_new = request.form.values()

            # cek kejadian errur halaman
            if csrf_token == None or csrf_token == "":
                raise ValueError("CSRF token tidak valid atau tidak ditemukan")
            if password_new != password2_new:
                raise ValueError("Password tidak sama")
            if email.endswith("@gmail.com") == False or email == "":
                raise ValueError("Email tidak valid")

            # eksekusi pasword
            password_hash = hashlib.sha256(password_new.encode()).hexdigest()
            result = db.users.find_one(
                {"email": email.lower(), "jobs": {"$in": ["Magang", "Karyawan"]}}
            )

            # cek lanjutan dan update
            if not result:
                raise Exception("Email tidak ditemukan / jobs bukan Magang/Karyawan")
            if result["password"] == password_hash:
                raise ValueError("Password sama")
            Otp = OtpPasswordGenerator(email.lower(), result["nama"])
            if not Otp:
                raise Exception("Terjadi kesalahan dalam pembuatan otp")
            jwt_otp = jwt.encode(
                {
                    "otp": Otp.otp,
                    "password_hash": password_hash,
                    "user": str(result["_id"]),
                    "exp": datetime.datetime.utcnow() + datetime.timedelta(minutes=5),
                },
                secretKey,
                algorithm="HS256",
            )
            return redirect(
                url_for(
                    "forgetPasswordOtp",
                    jwt_otp=string_to_uuid_like(jwt_otp),
                    status="berhasil",
                    msg="OTP relah dikrimkan ke email anda",
                )
            )

        except ValueError as e:
            return redirect(url_for("forgetPassword", msg=e.args[0]))
        except Exception as e:
            return redirect(url_for("forgetPassword", msg=e.args[0]))

    # request method get
    msg = request.args.get("msg")
    return render_template("forgetPassword.html", msg=msg)


# bagian otp lupa password
@app.route("/sign-in/forget/otp/<jwt_otp>", methods=["GET", "POST"])
def forgetPasswordOtp(jwt_otp):
    """
    Forget Password OTP
    -

    tags:
        - Autentikasi

    Halaman untuk menginputkan OTP yang dikirimkan ke email pengguna.
    Fungsi ini digunakan untuk menginputkan OTP yang dikirimkan ke email pengguna.
    OTP berupa angka yang dikirimkan ke email pengguna dan berlaku hanya dalam 5 menit saja.

    Args:
    - jwt_otp (string): Token yang berisi OTP, password hash, dan id pengguna.

    Returns:
    - Halaman input OTP jika request method GET.
    - Redirect ke halaman reset password jika OTP benar.
    - Redirect ke halaman lupa password jika OTP salah atau kadaluarsa.

    Raises:
    - Exception: OTP tidak valid atau kadaluarsa.
    """

    if request.method == "POST":
        try:
            # inisisasi data
            token = jwt_otp
            otp = request.json.get("otp")
            csrf_token = request.headers.get("X-CSRF-TOKEN")

            # cek kejadian errur halaman
            if csrf_token == None or csrf_token == "":
                raise ValueError("CSRF token tidak valid atau tidak ditemukan")
            if otp == None or otp == "":
                raise ValueError("OTP tidak boleh kosong")
            if token == None or token == "":
                raise ValueError("Token tidak boleh kosong")

            token = uuid_like_to_string(token)
            if not token:
                raise Exception("decode gagal")
            # terjemahkan jwt
            token_otp = jwt.decode(token, secretKey, algorithms="HS256")

            # cek otp yang dimasukkan sama dengan diberikan sama?
            if int(otp) != token_otp["otp"]:
                raise ValueError("OTP tidak sama")

            result = db.users.update_one(
                {
                    "_id": ObjectId(token_otp["user"]),
                },
                {"$set": {"password": token_otp["password_hash"]}},
            )

            # berhasil atau tidak
            if result.modified_count <= 0:
                raise Exception("Terjadi kesalahan Update Password")

            return (
                jsonify(
                    redirect="/sign-in",
                    msg="Password updated",
                    status="success",
                    title="Update Password",
                ),
                200,
            )

        # catch erronya
        except jwt.ExpiredSignatureError:
            return (
                jsonify(redirect=url_for("forgetPassword", msg="Token Kadaluarsa")),
                500,
            )
        except jwt.InvalidTokenError:
            return (
                jsonify(redirect=url_for("forgetPassword", msg="Token TIdak valid")),
                500,
            )
        except ValueError as e:
            return (
                jsonify(
                    redirect=url_for(
                        "forgetPasswordOtp", jwt_otp=jwt_otp, msg=e.args[0]
                    )
                ),
                500,
            )
        except Exception as e:
            return (
                jsonify(
                    redirect=url_for(
                        "forgetPasswordOtp", jwt_otp=jwt_otp, msg=e.args[0]
                    )
                ),
                500,
            )

    # request method get
    result = request.args.get("status")
    msg = request.args.get("msg")
    return render_template(
        "forgetPasswordOtp.html", token=jwt_otp, msg=msg, result=result
    )


# ballon_faq post
@app.route("/ask", methods=["POST"])
def ask():
    """
    Menangani pengiriman pertanyaan/kendala dari pengguna.

    Membutuhkan request JSON dengan field berikut:
    - name: Nama lengkap pengguna
    - email: Alamat email pengguna
    - jobs: Jabatan/posisi pengguna
    - departement: Departemen pengguna
    - kendala: Deskripsi masalah/pertanyaan

    Header yang dibutuhkan:
    - X-CSRF-TOKEN: Token CSRF yang valid

    Mengembalikan:
    - Sukses: Response JSON dengan status dan pesan berhasil
    - Error: Response JSON dengan detail error dan status code 400

    Contoh penggunaan:
    ```
    POST /ask
    Headers:
        X-CSRF-TOKEN: token123
    Body:
        {
            "name": "Budi Santoso",
            "email": "budi@email.com",
            "jobs": "Developer",
            "departement": "IT",
            "kendala": "Tidak bisa login ke sistem"
        }
    ```
    """
    try:
        csrf_token = request.headers.get("X-CSRF-TOKEN")
        name, email, jobs, departement, kendala = request.json.values()

        # cek csrf token valid atau tidak
        if csrf_token == "" or csrf_token == None:
            return make_response(
                jsonify({"status": "error", "msg": "csrfmu tidak ada"}), 400
            )

        # The above code is performing input validation for the variables `name`, `email`, `jobs`, and
        # `departement`. It checks if any of these variables are empty or None after stripping any
        # leading or trailing whitespaces. If any of the variables are empty or None, it raises an
        # Exception with a corresponding error message indicating that the respective field is empty
        # and prompts the user to enter the required information.
        if name.strip() == "" or name.strip() == None:
            raise Exception("Nama anda Kosong, silahkan masukkan nama anda")
        if email.strip() == "" or email.strip() == None:
            raise Exception("email anda Kosong, silahkan masukkan email anda")
        if jobs.strip() == "" or jobs.strip() == None:
            raise Exception("jobs anda Kosong, silahkan masukkan jobs anda")
        if departement.strip() == "" or departement.strip() == None:
            raise Exception(
                "departement anda Kosong, silahkan masukkan departement anda"
            )
        if kendala.strip() == "" or kendala.strip() == None:
            raise Exception("kendala anda Kosong, silahkan masukkan kendala anda")
        # melakukan pengiroman kendala ke gmail
        result_kirim_email = FaqGmailSender(
            email.strip(),
            name.strip(),
            jobs.strip(),
            departement=departement.strip(),
            kendala=kendala.strip(),
        )
        if not result_kirim_email:
            raise Exception("Gagal mengirim email")

        # tambahakan faq baru ke database
        add_faq_to_db = db.faq.insert_one(
            {
                "message_id": result_kirim_email.message_id,
                "no_ticket": result_kirim_email.uuid_ticket,
                "email": email.strip(),
                "name": name.strip(),
                "jobs": jobs.strip(),
                "departement": departement.strip(),
                "kendala": kendala.strip(),
                "status": "Pending",
            }
        )
        if not add_faq_to_db:
            raise Exception("Gagal menambahkan faq ticket ke database")

        # render sigin dengan status success
        return (
            jsonify(
                {
                    "redirect": url_for(
                        "signIn",
                        msg=Markup(
                            "Pengaduan anda dengan no ticket <b class='poppins-semibold'>#"
                            + result_kirim_email.uuid_ticket
                            + "</b> telah kami terima"
                        ),
                        status="success",
                        title="Pengaduan berhasil!",
                    )
                }
            ),
            200,
        )
    # The above code is handling exceptions related to JWT (JSON Web Token) authentication. It catches
    # different types of exceptions that can occur during JWT verification:
    except Exception as e:
        return jsonify({"redirect": url_for("signIn", msg=e.args[0])}), 500


# update my profile
@app.route("/myProfiles", methods=["GET", "POST"])
def myProfiles():
    """
    My Profile
    -

    tags:
        - Admin / Sub Admin
        - Magang / Karyawan

    Fungsi ini digunakan untuk melihat dan memperbarui profil pengguna.

    Jika metode permintaan adalah GET, fungsi ini akan menampilkan halaman profil pengguna.
    Jika metode permintaan adalah POST, fungsi ini akan memperbarui informasi profil pengguna.

    Cara Penggunaan:
    - Untuk melihat profil pengguna:
      1. Pastikan pengguna telah login dan memiliki token yang valid.
      2. Kirim permintaan GET ke endpoint "/myProfiles".
      3. Halaman profil pengguna akan ditampilkan.

    - Untuk memperbarui profil pengguna:
      1. Pastikan pengguna telah login dan memiliki token yang valid.
      2. Siapkan data yang akan diperbarui, seperti email, nama, dan jika perlu, foto profil.
      3. Kirim permintaan POST ke endpoint "/myProfiles" dengan menyertakan data yang diperbarui.
      4. Jika berhasil, profil pengguna akan diperbarui.

    Catatan:
    - Permintaan POST memerlukan token CSRF yang valid untuk mencegah serangan CSRF.
    - Jika ada kesalahan dalam memperbarui profil, pesan kesalahan akan diberikan.
    """

    if request.method == "POST":
        try:
            # The above code is attempting to retrieve the value of the "token_key" cookie from the
            # request object in a Python web application. It uses the `get` method on the `cookies`
            # attribute of the `request` object to access the value of the "token_key" cookie. If the
            # cookie is present in the request, its value will be stored in the `token_key` variable.
            token_key = request.cookies.get("token_key")
            if token_key == None or token_key == "":
                raise ValueError("Token tidak boleh kosong")
            # The code snippet provided is attempting to convert a UUID-like object `token_key` to a
            # string using the `uuid_like_to_string` function. If the resulting `token_key` string is
            # empty, it raises a `ValueError` with the message "decode token error".
            token_key = uuid_like_to_string(token_key)
            if not token_key:
                raise ValueError("decode token error")
            # The above Python code snippet is decoding a JSON Web Token (JWT) using the `jwt.decode`
            # function with the specified secret key and algorithm "HS256". It then checks if the
            # decoded payload's "role" value is not equal to 3. If the condition is met, it retrieves
            # values from the form data (csrf_token, email, nama, and possibly a profile picture file)
            # using `request.form.values()` and `request.files.get("profile-pic")`.
            payloads = jwt.decode(token_key, secretKey, algorithms=["HS256"])
            if payloads["role"] != 3:
                csrf_token, email, nama = request.form.values()
                gambar = request.files.get("profile-pic")
            else:
                gambar = request.files.get("profile-pic")
                csrf_token, email, nama, nik, tempat_lahir, tanggal_lahir = (
                    request.form.values()
                )
                tanggal_lahir = (
                    datetime.datetime.strptime(tanggal_lahir, "%Y-%m-%d")
                    .strftime("%d %B %Y")
                    .lower()
                )

            # The above code is checking if the variable `gambar` is falsy (e.g., None, empty string,
            # 0). If `gambar` is falsy, it retrieves the `photo_profile` field from a document in the
            # `users` collection in a MongoDB database based on the `_id` field provided in the
            # `payloads` dictionary.
            if not gambar:
                gambar = db.users.find_one(
                    {"_id": ObjectId(payloads["_id"])}, {"photo_profile": 1}
                )["photo_profile"]

            # cek csrf token
            if csrf_token == None or csrf_token == "":
                raise ValueError("CSRF token tidak valid atau tidak ditemukan")

            # The above Python code snippet is checking the type of the variable `gambar`. If `gambar`
            # is a string, it sets the `filepath_db` variable to the value of `gambar`. If `gambar` is
            # a `FileStorage` object (typically used for file uploads in Flask), it checks the file
            # extension to ensure it is either ".png", ".jpg", or ".jpeg". If the extension is not
            # valid, it redirects to a URL with a message indicating the extension is not allowed. If
            # the file extension is valid, it secures the filename, saves the file to a specified
            # folder
            if type(gambar) == str:
                filepath_db = gambar
            elif type(gambar) == FileStorage:

                # cek ectension gambar
                if gambar.filename.endswith((".png", ".jpg", ".jpeg")) == False:
                    return redirect(
                        url_for("myProfiles", msg="tidak ditemukan extension .png .jpg .jpeg")
                    )

                # Kirim ke Imgbb
                response = upload_to_imgbb(gambar, imgbb_api_key)
                print(response)
                if response["status"] == "success":
                    filepath_db = response["url"] + " " + response["filename"]
                else:
                    raise Exception("Gagal Upload Gambar ke Imgbb")
            else:
                raise Exception("Gambar tidak valid")

            # The above code is updating user information in a MongoDB database based on the role of
            # the user. If the user's role is 3, it updates the user's email, name, ID number, place
            # of birth, date of birth, and profile photo. If the user's role is not 3, it updates the
            # user's email, name, and profile photo. The code uses the `update_one` method from the
            # `db.users` collection in MongoDB to update the user information.
            if payloads["role"] == 3:
                result = db.users.update_one(
                    {
                        "_id": ObjectId(payloads["_id"]),
                    },
                    {
                        "$set": {
                            "email": email.lower(),
                            "nama": nama.title(),
                            "nik": int(nik),
                            "tempat_lahir": tempat_lahir.title(),
                            "tanggal_lahir": tanggal_lahir,
                            "photo_profile": filepath_db,
                        }
                    },
                )
            else:
                result = db.users.update_one(
                    {
                        "_id": ObjectId(payloads["_id"]),
                    },
                    {
                        "$set": {
                            "email": email.lower(),
                            "nama": nama.title(),
                            "photo_profile": filepath_db,
                        }
                    },
                )

            # The above Python code is checking the value of the variable `result`. If `result`
            # evaluates to `True`, it will redirect the user to the "myProfiles" route with a success
            # message "Update Profile Berhasil". If `result` is `False`, it will redirect the user to
            # the "myProfiles" route with a failure message "Update Profile Gagal".
            if result:
                return redirect(
                    url_for(
                        "myProfiles", msg="Update Profile Berhasil", status="success"
                    )
                )
            else:
                raise Exception("Update Profil Gagal")

        # The above code snippet is handling exceptions in a Python function. It is using a try-except
        # block to catch specific exceptions and then redirecting the user to the "signIn" route with
        # a specific message based on the type of exception caught.
        except jwt.ExpiredSignatureError:
            return redirect(url_for("signIn", msg="Sesi Kadaluarsa!"))
        except jwt.DecodeError:
            return redirect(url_for("signIn", msg="Anda telah logout"))
        except ValueError as e:
            return redirect(url_for("signIn", msg=e.args[0]))
        except Exception as e:
            return redirect(url_for("myProfiles", msg=e.args[0]))

    # get method
    # The above code snippet is written in Python and it seems to be part of a web application using a
    # framework like Flask or Django. It is trying to retrieve the values of "msg" and "status"
    # parameters from the request arguments. These values are typically passed in the URL query string
    # when a user makes a request to the server. The retrieved values are stored in the variables
    # `msg` and `status` respectively.
    msg = request.args.get("msg")
    status = request.args.get("status")
    try:
        # The above Python code is checking for the presence and validity of CSRF token and token key
        # retrieved from cookies in a web request. If either the CSRF token or token key is missing or
        # empty, it raises a `ValueError` with a corresponding error message indicating that the CSRF
        # token or token key is not valid or not found.
        token_key = request.cookies.get("token_key")
        csrf_token = request.cookies.get("csrf_token")
        if csrf_token == None or csrf_token == "":
            raise ValueError("CSRF token tidak valid atau tidak ditemukan")
        if token_key == None or token_key == "":
            raise ValueError("Token key tidak valid atau tidak ditemukan")

        # decode uuid
        token_key = uuid_like_to_string(token_key)
        csrf_token = uuid_like_to_string(csrf_token)
        if not csrf_token:
            return redirect(url_for("signIn", msg="CSRF Token kadaluarsa"))
        if not token_key:
            return redirect(url_for("signIn", msg="Token Key kadaluarsa"))

        # decode jwt
        payloads = jwt.decode(token_key, secretKey, algorithms=["HS256"])

        # cek data user
        data = db.users.find_one({"_id": ObjectId(payloads["_id"])})

        # tampilkan
        return render_template("myProfiles.html", data=data, msg=msg, status=status)

    # The above code is handling different exceptions that may occur in a Python application.
    except jwt.ExpiredSignatureError:
        return redirect(url_for("signIn", msg="Session kadaluarsa"))
    except jwt.DecodeError:
        return redirect(url_for("signIn", msg="Anda telah logout"))
    except ValueError as ve:
        return redirect(url_for("signIn", msg=ve.args[0]))
    except Exception as e:
        return redirect(url_for("signIn", msg=e.args[0]))


# dasboard magang get
@app.route("/dashboard", methods=["GET"])
def dashboard():
    """
    Halaman Dashboard Pengguna / admin
    ---
    tags:
      - Admin / Sub Admin
      - Magang / Karyawan
    description: Menampilkan dashboard berdasarkan peran admin dan data kehadiran.
    parameters:
      - name: result
        in: query
        type: string
        required: false
        description: Hasil operasi sebelumnya (misalnya, sukses atau gagal).
        example: "success"
      - name: msg
        in: query
        type: string
        required: false
        description: Pesan tambahan yang akan ditampilkan di dashboard.
        example: "Login berhasil"
      - name : token_key
        in: cookie
        type: string
        required: true
        description: Token key yang digunakan untuk autentikasi pengguna.
        example: "your_token_key"
      - name : csrf_token
        in: cookie
        type: string
        required: true
        description: CSRF token yang digunakan untuk keamanan pengguna.
        example: "your_csrf_token"
    responses:
      200:
        description: Dashboard berhasil dimuat.
      401:
        description: Pengguna tidak terautentikasi.
      500:
        description: Terjadi kesalahan internal server.
    """
    result = request.args.get("result")
    msg = request.args.get("msg")
    cookie = request.cookies.get("token_key")
    csrf_token = request.cookies.get("csrf_token")

    try:
        # cek ada cookie dan csrf token / tidak ?
        if not cookie:
            raise Exception("Cookie kadaluarsa")
        if not csrf_token:
            raise Exception("CSRF Token kadaluarsa")
        # convet uuid
        cookie = uuid_like_to_string(cookie)
        csrf_token = uuid_like_to_string(csrf_token)

        # convert jwt
        payload = jwt.decode(cookie, secretKey, algorithms=["HS256"])

        # ambil data
        data = db.users.find_one(
            {
                "_id": ObjectId(payload["_id"]),
                "jobs": payload["jobs"],
                "role": payload["role"],
            }
        )

        # cek role dan jobs yang didapat untuk render template
        # bagian magang / karyawan
        if data["role"] == 3:
            # ambiltanggal sekarang
            data["tanggal_sekarang"] = get_time_zone_now().strftime("%d %B %Y").lower()

            # ambil data table absen
            table_hadir = db.absen_magang.find_one(
                {
                    "user_id": ObjectId(payload["_id"]),
                    "tanggal_hadir": data["tanggal_sekarang"],
                }
            )

            # cek tanggal pelaksanan kerja
            waktu_sekarang = get_time_zone_now()
            waktu30menit = waktu_sekarang + datetime.timedelta(minutes=30)
            waktu_sekarang = waktu_sekarang.time()
            waktu_awal_kerja = datetime.datetime.strptime(
                data["waktu_awal_kerja"].replace(".", ":"), "%H:%M"
            ).time()
            waktu_akhir_kerja = datetime.datetime.strptime(
                data["waktu_akhir_kerja"].replace(".", ":"), "%H:%M"
            ).time()

            # customize button berdasarkan jam kerja ,sebelum, dan sesudah tanpa melakukan klik hadir
            if waktu_awal_kerja <= waktu_sekarang:
                # ini saat mulai klik hadir
                if waktu_sekarang <= waktu_akhir_kerja:
                    data["class-button-hadir"] = "btn-primary"
                    data["text-button"] = "Hadir"
                    # dah kelar tapi belum diklik bang
                else:
                    data["class-button-hadir"] = "btn-danger disabled opacity-100"
                    data["text-button"] = "Tidak Hadir"
            else:
                data["class-button-hadir"] = "btn-secondary disabled"
                data["text-button"] = "Belum Absen"

            # ambil anggka heading card dan customize button berdasarkan cek status hadir dan waktu
            if table_hadir:
                # konversi waktu hadir
                waktu_hadir = datetime.datetime.strptime(
                    table_hadir["waktu_hadir"].replace(".", ":"), "%H:%M"
                ).time()
                # membuat jumlah absen dalam rentang kontrak
                data["heading-card"] = (
                    data["absen"]["hadir"] + data["absen"]["tidak_hadir"]
                )
                # saat jam awal dan sudah dapat melakukan absen
                if (
                    table_hadir["status_hadir"] == "1"
                    and waktu_hadir <= waktu_akhir_kerja
                ):
                    data["class-button-hadir"] = "btn-success disabled opacity-100"
                    data["text-button"] = "Hadir"
                # saat jam akhir dan sudah dapat melakukan absen
                elif (
                    table_hadir["status_hadir"] == 2
                    and waktu_hadir <= waktu_akhir_kerja
                ):
                    data["class-button-hadir"] = "btn-warning disabled opacity-100"
                    data["text-button"] = "Telat"
                elif (
                    table_hadir["status_hadir"] in ("1", 2)
                    and waktu_hadir >= waktu_akhir_kerja
                ):
                    data["class-button-hadir"] = "btn-secondary disabled"
                    data["text-button"] = "Selesai"
                else:
                    pass
            else:
                data["heading-card"] = (
                    data["absen"]["hadir"] + data["absen"]["tidak_hadir"] + 1
                )
            print(data)
            # ini buat lihat hadir dan tidak hadir
            data2 = [
                {
                    "titleCard": "Hadir",
                    "textColor": "text-success",
                    "angkaCard": data["absen"]["hadir"],
                    "icon": "fas fa-check fa-2x text-success text-opacity-25",
                    "borderLeft": "border-left-success",
                },
                {
                    "titleCard": "Telat",
                    "textColor": "text-warning",
                    "angkaCard": data["absen"]["telat"],
                    "icon": "fa-solid fa-triangle-exclamation fa-2x text-warning text-opacity-25",
                    "borderLeft": "border-left-warning",
                },
                {
                    "titleCard": "Tidak Hadir",
                    "textColor": "text-danger",
                    "angkaCard": data["absen"]["tidak_hadir"],
                    "icon": "fa-regular fa-circle-xmark fa-2x text-danger text-opacity-25",
                    "borderLeft": "border-left-danger",
                },
                {
                    "titleCard": "Liburan",
                    "textColor": "text-danger-emphasis",
                    "angkaCard": data["absen"]["libur"],
                    "icon": "fa-solid fa-bed fa-2x text-danger-emphasis text-opacity-25",
                    "borderLeft": "border-left-danger-emphasis",
                },
            ]
            # hanya untuk karywan dan magang
            if (data["jobs"] == "Magang") or (data["jobs"] == "Karyawan"):
                return render_template(
                    "dashboard_magang.html",
                    data=data,
                    data2=data2,
                    msg=msg,
                    result=result,
                )
            else:
                raise Exception("Anda bukan Magang / Karyawan")
        # bagian admin
        elif data["role"] == 1:
            if data["jobs"] == "Admin" or data["jobs"] == "Sub Admin":
                if data['departement'] == 'Mentor' and data['jobs'] == 'Sub Admin' :
                    return redirect(url_for('task'))
                # panjang table awal
                first_table = request.args.get("page_awal", 1, type=int)
                # panjang table
                table_length = request.args.get("length", 10, type=int)
                # Hitung offset
                skip = (first_table - 1) * table_length
                # atur search
                search = request.args.get("search", default="", type=str)
                data_karyawan_magang_new = []

                # ambil data limit
                data_karyawan_magang = list(
                    db.users.find({"jobs": {"$in": ["Magang", "Karyawan"]}})
                    .skip(skip)
                    .limit(table_length)
                )

                # hitunhg data
                count_data_karyawan_magang = db.users.count_documents(
                    {"jobs": {"$in": ["Magang", "Karyawan"]}}
                )
                # Hitung showing
                if count_data_karyawan_magang == 0:
                    start_index = skip
                else:
                    start_index = skip + 1

                end_index = min(skip + table_length, count_data_karyawan_magang)

                # tambah data indeks
                data["total_page"] = (
                    count_data_karyawan_magang + table_length - 1
                ) // table_length
                data["first_table"] = first_table  # table awal
                data["table_length"] = table_length  # table akhir
                data["count_data_karyawan_magang"] = count_data_karyawan_magang

                # fitur pencarian
                if search != "" and search != None:
                    for data_karyawan in data_karyawan_magang:
                        # cek berdasarkan jobs,role,departement,nama,nik,email,tanggal_lahir,tempat_lahir
                        if (
                            search in data_karyawan["jobs"].lower()
                            or search in str(data_karyawan["role"])
                            or search in data_karyawan["departement"].lower()
                            or search in data_karyawan["nama"].lower()
                            or search in str(data_karyawan["nik"])
                            or search in data_karyawan["email"].lower()
                            or search in data_karyawan["tempat_lahir"].lower()
                            or search in data_karyawan["tanggal_lahir"].lower()
                        ):
                            data_karyawan_magang_new.append(data_karyawan)
                        continue
                    # cek datakaryawan ditemukan atau tidak
                    if len(data_karyawan_magang_new) <= 0:
                        return redirect(
                            url_for("dashboard", msg="Data tidak ditemukan")
                        )
                    # ditemukan
                    data_karyawan_magang = data_karyawan_magang_new
                    end_index = len(data_karyawan_magang)
                # editable showing
                data["showing_table_karyawan"] = (
                    f"Showing {start_index} to {end_index} of {count_data_karyawan_magang} entries"
                )
                # judul
                data["heading_title_body"] = "Dashboard"

                # render
                return render_template(
                    "dashboard_magang.html",
                    data=data,
                    data_karyawan_magang=data_karyawan_magang,
                    msg=msg,
                    search=search,
                    result=result,
                )
            else:
                raise Exception("Anda bukan administration")

        else:
            raise Exception("Anda bukan bagian dari perusahaan")
    # handle error
    except jwt.ExpiredSignatureError:
        return redirect(url_for("signIn", msg="Session kadaluarsa"))
    except jwt.DecodeError:
        return redirect(url_for("signIn", msg="Anda telah logout"))
    except Exception as e:
        return redirect(url_for("signIn", msg=e.args[0]))


# dashboard absen
@app.route("/dashboard/absen", methods=["POST"])
def dashboardAbsen():
    """
    Catat Kehadiran
    ---
    tags:
      - Magang / Karyawan
    description: Mengizinkan pengguna untuk mencatat kehadiran mereka.
    consumes:
      - "application/x-www-form-urlencoded"
    parameters:
      - name: user_id
        in: formData
        type: string
        required: true
        description: ID pengguna.
      - name: status_hadir
        in: formData
        type: string
        required: true
        description: Status hadir pengguna.
      - name: "X-CSRF-Token"
        in: header
        required: true
        type: string
        description: CSRF token.
      - name: "csrf_token"
        in: cookie
        type: string
        required: true
        description: CSRF token Cookie.
    responses:
      200:
        description: Kehadiran berhasil dicatat.
      400:
        description: Input tidak valid.
      500:
        description: Terjadi kesalahan internal server.
    """

    # ambil tanggal sekarang dan waktu sekarang
    now = get_time_zone_now()
    time_now = now.time()
    try:
        # ambil csrf token
        csrf_token = request.headers.get("X-CSRF-Token")
        csrf = request.cookies.get("csrf_token")
        # cek token
        if csrf_token == "" and csrf_token == None:
            raise Exception("csrf token Kadaluarsa")
        if csrf == "" and csrf == None:
            raise Exception("csrf token cookie Kadaluarsa")
        csrf = uuid_like_to_string(csrf)
        if not csrf:
            raise Exception("CSRF decode error")
        # ambil data dari form
        userId = request.form.get("user_id")
        status_hadir = request.form.get("status_hadir")

        # ambil riwayat absen
        riwayat_absen = db.absen_magang.find_one(
            {"user_id": ObjectId(userId)}, sort={"_id": -1}
        )

        # cek sudah klik button absen / belum sebelumnya
        if riwayat_absen:
            if (
                now.date()
                <= datetime.datetime.strptime(
                    riwayat_absen["tanggal_hadir"], "%d %B %Y"
                ).date()
                and time_now
                >= datetime.datetime.strptime(riwayat_absen["waktu_hadir"], "%H.%M").time()
            ):
                return (
                    jsonify(
                        {
                            "result": "unsuccess",
                            "redirect": url_for(
                                "dashboard",
                                msg="Anda sudah absen pada tanggal "
                                + riwayat_absen["tanggal_hadir"],
                            ),
                        }
                    ),
                    200,
                )
        # cek status hadir bukan 1 ubah jadi int
        if status_hadir != "1":
            status_hadir = int(status_hadir)

        # jika hadir
        if status_hadir == "1":
            setting = {
                "absen.hadir": db.users.find_one({"_id": ObjectId(userId)})["absen"][
                    "hadir"
                ]
                + 1
            }
        # jika telat
        else:
            setting = {
                "absen.telat": db.users.find_one({"_id": ObjectId(userId)})["absen"][
                    "telat"
                ]
                + 1
            }

        # lakukan update
        result = db.users.find_one_and_update(
            {"_id": ObjectId(userId)},
            {"$set": setting},
        )
        if not result:
            raise Exception("Terjadi kesalahan dalam update jumlah absen ke db")

        # lakukan penambahan absen magang riwayat
        result = db.absen_magang.insert_one(
            {
                "user_id": ObjectId(userId),
                "status_hadir": status_hadir,
                "waktu_hadir": get_time_zone_now().strftime("%H.%M").lower(),
                "tanggal_hadir": get_time_zone_now().strftime("%d %B %Y").lower(),
            }
        )
        if not result:
            raise Exception("Terjadi kesalahan dalam penambahan riwayat ke db")
        # success
        return jsonify({"result": "success", "redirect": "/dashboard"}), 200
    # handle error
    except jwt.ExpiredSignatureError:
        return jsonify({"redirect": url_for("signIn", msg="Session Kadaluarsa")}), 500
    except jwt.DecodeError:
        return jsonify({"redirect": url_for("signIn", msg="Anda telah logout")}), 500
    # except Exception as e:
    #     if not e.args:
    #         e.args = [
    #             "terjadi kesalahan data",
    #         ]
    #     return jsonify({"redirect": url_for("signIn", msg=f"{e.args[0]}")}), 500


# change password
@app.route("/change-password", methods=["GET", "POST"])
def change_password():
    """
    Ubah Password
    -
    tags:
        - admin / sub admin
        - magang / karyawan

    Endpoint ini digunakan untuk mengubah password pengguna.

    Parameters:
        - old_password (string): Password lama pengguna.
        - new_password (string): Password baru pengguna.
        - confirm_password (string): Konfirmasi password baru pengguna.

    Returns:
        - Redirect ke halaman login dengan pesan sukses jika berhasil.
        - Redirect ke halaman login dengan pesan gagal jika gagal.
    """

    # method post
    if request.method == "POST":
        try:
            # ambil dulu cookie dan headers
            cookie = request.cookies.get("token_key")
            csrf_token = request.cookies.get("csrf_token")
            X_CSRFToken = request.headers.get("X-CSRFToken")
            # cek csrf token
            if csrf_token == None or csrf_token == "":
                return jsonify(
                    {
                        "status": 400,
                        "redirect": url_for("sign-in", msg="CSRF Token Kadaluarsa"),
                    }
                )
            # cek xcsrf token
            if X_CSRFToken == None or X_CSRFToken == "":
                return jsonify(
                    {
                        "status": 400,
                        "redirect": url_for("sign-in", msg="X-CSRF Token Kadaluarsa"),
                    }
                )
            # cek cookie
            if cookie == None or cookie == "":
                return jsonify(
                    {
                        "status": 400,
                        "redirect": url_for("sign-in", msg="Cookie Kadaluarsa"),
                    }
                )

            # ambil dulu convert
            cookie = uuid_like_to_string(cookie)
            csrf_token = uuid_like_to_string(csrf_token)
            if not csrf_token:
                return jsonify(
                    {
                        "status": 401,
                        "redirect": url_for("sign-in", msg="CSRF Token Kadaluarsa"),
                    }
                )
            if not cookie:
                return jsonify(
                    {
                        "status": 401,
                        "redirect": url_for("sign-in", msg="Cookie Kadaluarsa"),
                    }
                )

            # decode payload
            payload = jwt.decode(cookie, secretKey, algorithms=["HS256"])
            if payload["role"] not in (1, 3):
                raise Exception("anda tidak punya hak akses")

            # ambil data password di database
            password_db = db.users.find_one(
                {
                    "_id": ObjectId(payload["_id"]),
                    "jobs": payload["jobs"],
                    "role": payload["role"],
                    "password": hashlib.sha256(
                        request.form.get("old_password").encode()
                    ).hexdigest(),
                },
                {"_id": 0, "password": 1},
            )
            # cek sama atau tidak
            if not password_db:
                raise Exception("password lama salah")

            # cek password baru dengan lama
            if (
                hashlib.sha256(request.form.get("new_password").encode()).hexdigest()
                == password_db["password"]
            ):
                raise Exception("password baru tidak boleh sama dengan lama")

            # update password
            result = db.users.update_one(
                {
                    "_id": ObjectId(payload["_id"]),
                    "jobs": payload["jobs"],
                    "role": payload["role"],
                },
                {
                    "$set": {
                        "password": hashlib.sha256(
                            request.form.get("new_password").encode()
                        ).hexdigest()
                    }
                },
            )

            # cek berhasil atau tidak
            if result.modified_count <= 0:
                raise Exception("Gagal merubah password")
            return jsonify(
                {
                    "status": 200,
                    "redirect": url_for(
                        "change_password",
                        msg="Password Berhasil DIubah",
                        result="success",
                    ),
                }
            )
        # atasi error jwt
        except jwt.ExpiredSignatureError:
            return jsonify(
                {"status": 401, "redirect": url_for("signIn", msg="Session Kadaluarsa")}
            )
        except jwt.DecodeError:
            return jsonify(
                {"status": 408, "redirect": url_for("signIn", msg="Anda telah logout")}
            )
        # atasi error lainnya
        except Exception as e:
            return jsonify(
                {"status": 400, "redirect": url_for("change_password", msg=e.args[0])}
            )

    # method get
    try:
        # ambil data
        cookie = request.cookies.get("token_key")
        csrf_token = request.cookies.get("csrf_token")
        result = request.args.get("result")
        msg = request.args.get("msg")
        # cek ada ga?
        if csrf_token == None or csrf_token == "":
            return redirect(url_for("signIn", msg="csrf token Kadaluarsa"))
        if cookie == None or cookie == "":
            return redirect(url_for("signIn", msg="Cookie Kadaluarsa"))
        # konvert
        cookie = uuid_like_to_string(cookie)
        csrf_token = uuid_like_to_string(csrf_token)
        # konvert ke 2
        payload = jwt.decode(cookie, secretKey, algorithms=["HS256"])
        data = db.users.find_one(
            {
                "_id": ObjectId(payload["_id"]),
                "jobs": payload["jobs"],
                "role": payload["role"],
            }
        )
        data["heading_title_body"] = "Change Password"
        return (
            render_template("change_password.html", data=data, result=result, msg=msg),
            200,
        )
    except jwt.ExpiredSignatureError:
        return redirect(url_for("signIn", msg="Session Kadaluarsa"))
    except jwt.DecodeError:
        return redirect(url_for("signIn", msg="Anda telah logout"))


# riwayat kehadiran
@app.route("/riwayat-kehadiran", methods=["GET"])
def riwayat_kehadiran():
    """
    Tambah Riwayat Kehadiran
    ---
    tags:
      - Admin / Sub Admin
      - Magang / Karyawan
    summary: Tambahkan Riwayat Kehadiran
    description: Menambahkan data kehadiran baru untuk pengguna.
    parameters:
      - name: msg
        in: query
        type: string
        required: false
        description: pesan message error atau success.
        example: 'ini pesan message'
      - name: result
        in: query
        type: string
        required: false
        description: status pesan dari alert .
        example: 'success'
      - name : token_key
        in: cookie
        type: string
        required: true
        description: Token key yang digunakan untuk autentikasi pengguna.
        example: "your_token_key"
      - name : csrf_token
        in: cookie
        type: string
        required: true
        description: CSRF token yang digunakan untuk keamanan pengguna.
        example: "your_csrf_token"
    responses:
      200:
        description: Kehadiran berhasil ditambahkan.
      400:
        description: Input tidak valid.
      500:
        description: Terjadi kesalahan internal server.
    """

    try:
        # The above Python code is checking for the presence of a CSRF token and a token key in the
        # request cookies. If the CSRF token is not found (i.e., None), it redirects the user to the
        # "signIn" route with a message indicating that the CSRF token has expired. Similarly, if the
        # token key is not found (i.e., None), it redirects the user to the "signIn" route with a
        # message indicating that the user has been logged out. This code is likely part of a web
        # application's security mechanism to ensure that valid CSRF tokens and user authentication
        # tokens are present before allowing access
        msg = request.args.get("msg")
        result = request.args.get("result")
        csrf_token = request.cookies.get("csrf_token")
        cookies = request.cookies.get("token_key")
        if csrf_token == None:
            return redirect(url_for("signIn", msg="csrf token Kadaluarsa"))
        if cookies == None:
            return redirect(url_for("signIn", msg="Anda Telah logout"))
        # The above Python code snippet is converting `cookies` and `csrf_token` variables from a
        # UUID-like format to a string format using the `uuid_like_to_string` function. It then checks
        # if the `csrf_token` or `cookies` are empty, and if so, it redirects the user to the "signIn"
        # route with a message indicating that either the CSRF token has expired or the cookie has
        # expired.
        cookies = uuid_like_to_string(cookies)
        csrf_token = uuid_like_to_string(csrf_token)
        if not csrf_token:
            return redirect(url_for("signIn", msg="CSRF Token Kadaluarsa"))
        if not cookies:
            return redirect(url_for("signIn", msg="Cookie Kadaluarsa"))
        # The above Python code snippet is decoding a JSON Web Token (JWT) from the `cookies` using a
        # `secretKey` with the HS256 algorithm. It then retrieves data from a MongoDB database based
        # on the decoded payload. If the role in the payload is not equal to 1 and the job is not
        # 'admin', it fetches the attendance history (`riwayat_absent`) for a specific user ID.
        # Otherwise, if the conditions are not met, it fetches the attendance history for all users.
        payloads = jwt.decode(cookies, secretKey, algorithms=["HS256"])
        data = db.users.find_one(
            {"_id": ObjectId(payloads["_id"])},
            {
                "_id": 0,
                "jobs": 1,
                "role": 1,
                "nama": 1,
                "departement": 1,
                "photo_profile": 1,
            },
        )
        if not data:
            return redirect(url_for("dashboard", msg="Terjadi kesalahan data"))
        if payloads["role"] != 1 and payloads["jobs"] not in ("Admin", "Sub Admin"):
            riwayat_absent = list(
                db.absen_magang.find(
                    {"user_id": ObjectId(payloads["_id"])},
                    {"_id": 0, "status_hadir": 1, "waktu_hadir": 1, "tanggal_hadir": 1},
                )
            )
        else:
            # gabungkan data riwayat absen karyawan dan magang
            riwayat_absent = list(
                db.absen_magang.aggregate(
                    [
                        {
                            "$lookup": {
                                "from": "users",
                                "localField": "user_id",
                                "foreignField": "_id",
                                "as": "user",
                            }
                        },
                        {"$unwind": "$user"},
                        {
                            "$match": {
                                "user.role": {"$ne": 1},
                                "user.jobs": {"$nin": ["admin"]},
                            }
                        },
                        {"$unset": ["user_id", "user._id", "user.password"]},
                        {
                            "$project": {
                                "_id": 1,
                                "status_hadir": 1,
                                "waktu_hadir": 1,
                                "tanggal_hadir": 1,
                                "user.nama": 1,  # Tampilkan `nama` dari `users`
                                "user.jobs": 1,  # Tampilkan `jobs` dari `users`
                                "user.departement": 1,  # Tampilkan `departement` dari `users`
                                "user.role": 1,  # Tampilkan `role` dari `users`
                                "user.email": 1,
                                "user.nik": 1,
                            }
                        },
                    ]
                )
            )
            if riwayat_absent:
                for i in riwayat_absent:
                    # decrypt id riwayat absen
                    i["_id"] = string_to_uuid_like(
                        str(i["_id"])  
                    )

        # render template riwayat_kehadiran.html
        return (
            render_template(
                "riwayat_kehadiran.html",
                riwayat_absent=riwayat_absent,
                data=data,
                msg=msg,
                result=result,
            ),
            200,
        )

    # The above code is a Python snippet that includes multiple `except` blocks to handle different
    # types of exceptions.
    except jwt.ExpiredSignatureError:
        return redirect(url_for("signIn", msg="Session Kadaluarsa"))
    except jwt.DecodeError:
        return redirect(url_for("signIn", msg="Anda telah logout"))
    except Exception:
        return redirect(url_for("signIn", msg="something went wrong!"))


# riwayat kehadiran post
# lanjutkan disini
@app.route("/riwayat-kehadiran/<path1>", methods=["POST"])
@app.route("/riwayat-kehadiran/<path1>/<path2>", methods=["POST"])
def riwayat_kehadiran_post(path1=None, path2=None):
    """
    Fungsi untuk mengelola riwayat kehadiran
    =

    Fungsi ini digunakan untuk mengelola riwayat kehadiran karyawan dan magang.

    Parameters
    -

    * path1 : str
        Path pertama yang digunakan untuk mengelola riwayat kehadiran. Nilai yang valid adalah "edit" dan "delete".
    * path2 : str, optional
        Path kedua yang digunakan untuk mengelola riwayat kehadiran. Nilai yang valid adalah id riwayat kehadiran.

    Returns
    -

    * redirect : str
        Redirect ke halaman riwayat kehadiran dengan pesan.

    Notes
    -

    * Fungsi ini hanya dapat diakses oleh admin dan sub admin.
    * Fungsi ini menggunakan JWT untuk validasi cookie.
    * Fungsi ini menggunakan MongoDB untuk menyimpan data riwayat kehadiran.
    * Fungsi ini menggunakan Flask untuk membuat aplikasi web.
    """

    try:
        # lakukan validasi cooke masih ada
        csrf_token = request.cookies.get("csrf_token")
        cookies = request.cookies.get("token_key")
        print(csrf_token, cookies)
        if csrf_token == None:
            return redirect(url_for("signIn", msg="csrf token Kadaluarsa"))
        if cookies == None:
            return redirect(url_for("signIn", msg="Anda Telah logout"))

        # lakukan decode uuid validasi cookie
        cookies = uuid_like_to_string(cookies)
        csrf_token = uuid_like_to_string(csrf_token)
        if not csrf_token:
            return redirect(url_for("signIn", msg="CSRF Token Kadaluarsa"))
        if not cookies:
            return redirect(url_for("signIn", msg="Cookie Kadaluarsa"))

        # lakukan encode payloads
        payloads = jwt.decode(cookies, secretKey, algorithms=["HS256"])
        if not payloads:
            return redirect(url_for("signIn", msg="Anda Telah logout"))

        # jika pathnya edit
        if (
            path1 == "edit"
            and payloads["role"] == 1
            and payloads["jobs"] in ("Admin", "Sub Admin")
        ):
            # method put
            if request.form["__method"] == "PUT":
                # form csrf dimiliki
                if (
                    request.form["__csrf_token"] != None
                    or request.form["__csrf_token"] != ""
                ):

                    # inisiasi form
                    id_riwayat_absent = request.form["__id_riwayat_absent"]
                    nik = request.form["nik"]
                    email = request.form["email"]
                    status_hadir = request.form["status_hadir"]

                    # cek status hadir dimiliki dan konversi
                    if status_hadir == None or status_hadir == "":
                        raise Exception("Anda harus memilih status hadir")
                    elif status_hadir != "1":
                        status_hadir = int(status_hadir)

                    # lakukan validasi nik dimikii
                    if nik == None or nik == "":
                        raise Exception("Anda harus mengisi nik profile terlebih dahulu")
                    
                    nik = int(nik)

                    # lakukan validasi email dan id_riwayat_absent
                    if "" in (email, id_riwayat_absent) or None in (
                        email,
                        id_riwayat_absent,
                    ):
                        raise Exception("This data is undefined please try again")
                    

                    # decript id riwayat absen
                    absen_magang_id = uuid_like_to_string(id_riwayat_absent)
                    # print(absen_magang_id, type(absen_magang_id))

                    # do change it absen_magang
                    update_absen_magang = db.absen_magang.find_one_and_update(
                        {"_id": ObjectId(absen_magang_id)},
                        {"$set": {"status_hadir": status_hadir}},
                    )

                    # cek berhasil diupdate
                    if not update_absen_magang:
                        raise Exception("Gagal update status absent")

                    # lakukan kondisi untuk perubahan status_hadir referrer user
                    if update_absen_magang["status_hadir"] == "1":
                        setting = {
                            "$set": {
                                "absen.hadir": db.users.find_one(
                                    {"_id": update_absen_magang["user_id"]}
                                )["absen"]["hadir"]
                                - 1
                            }
                        }
                    elif update_absen_magang["status_hadir"] == 2:
                        setting = {
                            "$set": {
                                "absen.telat": db.users.find_one(
                                    {"_id": update_absen_magang["user_id"]}
                                )["absen"]["telat"]
                                - 1
                            }
                        }
                    elif update_absen_magang["status_hadir"] == 3:
                        setting = {
                            "$set": {
                                "absen.libur": db.users.find_one(
                                    {"_id": update_absen_magang["user_id"]}
                                )["absen"]["libur"]
                                - 1
                            }
                        }
                    else:
                        setting = {
                            "$set": {
                                "absen.tidak_hadir": db.users.find_one(
                                    {"_id": update_absen_magang["user_id"]}
                                )["absen"]["tidak_hadir"]
                                - 1
                            }
                        }

                    # tambahkan seting absen pada user untuk status
                    if status_hadir == "1":
                        setting["$set"]["absen.hadir"] = (
                            db.users.find_one({"_id": update_absen_magang["user_id"]})[
                                "absen"
                            ]["hadir"]
                            + 1
                        )
                    elif status_hadir == 2:
                        setting["$set"]["absen.telat"] = (
                            db.users.find_one({"_id": update_absen_magang["user_id"]})[
                                "absen"
                            ]["telat"]
                            + 1
                        )
                    elif status_hadir == 3:
                        setting["$set"]["absen.libur"] = (
                            db.users.find_one({"_id": update_absen_magang["user_id"]})[
                                "absen"
                            ]["libur"]
                            + 1
                        )
                    else:
                        setting["$set"]["absen.tidak_hadir"] = (
                            db.users.find_one({"_id": update_absen_magang["user_id"]})[
                                "absen"
                            ]["tidak_hadir"]
                            + 1
                        )

                    # laukan perubahan absen pada user
                    update_absen_user = db.users.find_one_and_update(
                        {"_id": update_absen_magang["user_id"]}, setting
                    )

                    # cek berhasil melakukan update absen pada user
                    if not update_absen_user:
                        raise Exception("Gagal update absent bagian user")

                    return redirect(
                        url_for(
                            "riwayat_kehadiran",
                            msg=f"Riwayat absen atas nama {update_absen_user['nama']} berhasil diupdate",
                            result="success",
                        )
                    )

                else:
                    raise Exception("Please do with page web")
            else:
                raise Exception("This method is not allowed")

        # delete kehadiran kryawan / magang
        elif (
            path1 == "delete"
            and payloads["role"] == 1
            and payloads["jobs"] in ("Admin", "Sub Admin")
        ):
            if request.form["_method"] == "DELETE":
                # form csrf dimiliki
                if (
                    request.form["__csrf_token"] != None
                    or request.form["__csrf_token"] != ""
                ):
                    # inisiasi form yang diperlukan
                    if path2 == None or path2 == "":
                        raise Exception("This data is undefined please try again")

                    # decript id riwayat absen
                    absen_magang_id = uuid_like_to_string(path2)  

                    # lakukan validasi email dan id_riwayat_absent
                    if not absen_magang_id or absen_magang_id == "":
                        raise Exception(
                            "This data decrypt is undefined please try again"
                        )

                    # cari riwayat absen
                    riwayat_absen = db.absen_magang.find_one_and_delete(
                        {"_id": ObjectId(absen_magang_id)}
                    )

                    # lakukan pengurangan bagian user sesuai id riwayat absen
                    if not riwayat_absen:
                        raise Exception("Riwayat absen tidak ditemukan")

                    # cari user berdasarkan id riwayat absen
                    user = db.users.find_one({"_id": riwayat_absen["user_id"]})

                    # cek user didapatkan atau tidak
                    if not user:
                        raise Exception("User tidak ditemukan")

                    # cek status hadir riwayat absen dalam bentuk angka dan string
                    if riwayat_absen["status_hadir"] == "1":
                        setting = {"$set": {"absen.hadir": user["absen"]["hadir"] - 1}}
                    elif riwayat_absen["status_hadir"] == 2:
                        setting = {"$set": {"absen.telat": user["absen"]["telat"] - 1}}
                    elif riwayat_absen["status_hadir"] == 3:
                        setting = {"$set": {"absen.libur": user["absen"]["libur"] - 1}}
                    else:
                        setting = {
                            "$set": {
                                "absen.tidak_hadir": user["absen"]["tidak_hadir"] - 1
                            }
                        }

                    # lakukan pengurangan pada user
                    delete_absen_user = db.users.find_one_and_update(
                        {"_id": riwayat_absen["user_id"]}, setting
                    )

                    # cek user didapatkan atau tidak
                    if not delete_absen_user:
                        raise Exception("Gagal update absent bagian user")

                    return redirect(
                        url_for(
                            "riwayat_kehadiran",
                            msg=f"Riwayat absen atas nama {delete_absen_user['nama']} berhasil di hapus",
                            result="success",
                        )
                    )
                # error csrf
                raise Exception("Please do with page web")
            # error custom method
            else:
                raise Exception("This method is not allowed")
        # error auth user
        else:
            return redirect(
                url_for("signin", msg="Anda tidak memiliki akses terhadap halaman ini")
            )

    # handle error
    except jwt.ExpiredSignatureError:
        return redirect(url_for("signIn", msg="Session Kadaluarsa"))
    except jwt.DecodeError:
        return redirect(url_for("signIn", msg="Anda telah logout"))
    except Exception as e:
        if not e.args:
            e.args = ("terjadi kesalahan data",)
        return redirect(url_for("riwayat_kehadiran", msg=e.args[0]))


# riwayat bantuan
@app.route("/riwayat-bantuan", methods=["GET"])
def riwayat_bantuan():
    """
    Fungsi riwayat_bantuan
    -

    Fungsi ini digunakan untuk menampilkan riwayat bantuan dalam aplikasi.
    Fungsi ini akan memeriksa keberadaan dan validitas token CSRF dan cookie dalam permintaan web.
    Jika token atau cookie tidak valid, pengguna akan diarahkan kembali ke halaman masuk dengan pesan yang sesuai.

    Parameter:
    -
    - Tidak ada parameter yang diterima oleh fungsi ini.

    Proses:-
    1. Mendapatkan token CSRF dan cookie dari permintaan.
    2. Memeriksa apakah token CSRF dan cookie ada dan valid.
    3. Jika tidak valid, mengarahkan pengguna kembali ke halaman masuk.
    4. Mendekode payload dari cookie untuk mendapatkan informasi pengguna.
    5. Memeriksa peran (role) pengguna, jika bukan admin, mengarahkan ke dashboard dengan pesan akses ditolak.
    6. Mengambil data pengguna dari basis data berdasarkan ID yang terdapat dalam payload cookie.

    Pengembalian:
    -
    - Tidak ada pengembalian data secara langsung, tetapi mengarahkan pengguna ke halaman lain berdasarkan hasil pemeriksaan token dan cookie.
    """

    # Function implementation continues...
    # The above Python code snippet is checking for the presence and validity of a CSRF token and a
    # cookie in a web request. Here is a breakdown of the code:
    cookie = request.cookies.get("token_key")
    csrf_token = request.cookies.get("csrf_token")
    if csrf_token == None:
        return redirect(url_for("signIn", msg="csrf token Kadaluarsa"))
    if cookie == None:
        return redirect(url_for("signIn", msg="Anda Telah logout"))

    # convert this to data real and check if success or not
    cookie = uuid_like_to_string(cookie)
    csrf_token = uuid_like_to_string(csrf_token)
    if not csrf_token:
        return redirect(url_for("signIn", msg="CSRF Token Kadaluarsa"))
    if not cookie:
        return redirect(url_for("signIn", msg="Cookie Kadaluarsa"))

    try:
        status = request.args.get("status")
        msg = request.args.get("msg")

        # decode payload
        payloads = jwt.decode(cookie, secretKey, algorithms=["HS256"])
        if payloads["role"] != 1:
            return redirect(url_for("dashboard", msg="Anda tidak memiliki akses"))
        data = db.users.find_one(
            {"_id": ObjectId(payloads["_id"])},
            {
                "_id": 0,
                "jobs": 1,
                "role": 1,
                "nama": 1,
                "departement": 1,
                "photo_profile": 1,
            },
        )
        if not data:
            return redirect(url_for("dashboard", msg="Terjadi kesalahan data"))
        if not (data['departement']=='Mentor' and data['jobs']=='Admin'):
            return redirect(url_for("dashboard", msg="Anda tidak memiliki akses"))
        # take data all
        data_faq_all = list(db.faq.find({}))
        if data_faq_all:
            for i in data_faq_all:
                i["_id"] = string_to_uuid_like(
                    str(i["_id"])  
                )
                if not i["_id"]:
                    raise Exception("Terjadi Kesalahan Data dalam encrypt")
                if i["status"] == "Pending":
                    i["selectClass"] = (
                        "border-warning border-2 rounded p-2 bg-transparent text-warning"
                    )
                elif i["status"] == "Diproses":
                    i["selectClass"] = (
                        "border-info border-2 rounded p-2 bg-transparent text-info"
                    )
                elif i["status"] == "Selesai":
                    i["selectClass"] = (
                        "border-success border-2 rounded p-2 bg-transparent text-success"
                    )
        return render_template(
            "riwayat_bantuan.html",
            data=data,
            riwayat_bantuan=data_faq_all,
            result=status,
            msg=msg,
        )

    # The above code is handling exceptions related to JWT (JSON Web Token) authentication. It catches
    # different types of exceptions that can occur during JWT verification:
    except jwt.ExpiredSignatureError:
        return redirect(url_for("signIn", msg="Session Kadaluarsa"))
    except jwt.DecodeError:
        return redirect(url_for("signIn", msg="Anda telah logout"))
    except IndexError as e:
        return redirect(url_for("riwayat_bantuan", msg=e.args[0]))
    except Exception as e:
        return redirect(url_for("riwayat_bantuan", msg=e.args[0]))


# update status bantuan
@app.route("/update-status-bantuan", methods=["POST"])
def riwayat_bantuan_post():
    """
    Dokumentasi riwayat_bantuan_post
    -
    Fungsi ini digunakan untuk memperbarui status bantuan yang telah diajukan pengguna.
    Fungsi ini menerima parameter berupa 'status' dan 'status_id' yang dikirimkan melalui request form-data.

    Parameter:

    - status: string, status yang akan di update (Diproses, Selesai, Ditolak)
    - status_id: int, id dari riwayat bantuan yang akan di update

    Return:
    - redirect: string, url yang akan di redirect ke halaman riwayat bantuan
    - 500: jika terjadi kesalahan internal server
    """

    # The above Python code snippet is checking for the presence and validity of a CSRF token and a
    # cookie in a web request. Here is a breakdown of the code:
    cookie = request.cookies.get("token_key")
    csrf_token = request.cookies.get("csrf_token")
    if csrf_token == None:
        return jsonify({"redirect": url_for("signIn", msg="csrf token Kadaluarsa")}), 500
    if cookie == None:
        return jsonify({"redirect": url_for("signIn", msg="Anda Telah logout")}), 500

    # convert this to data real and check if success or not
    cookie = uuid_like_to_string(cookie)
    csrf_token = uuid_like_to_string(csrf_token)
    if not csrf_token or not cookie:
        return (
            jsonify({"redirect": url_for("signIn", msg="convert token failed!")}),
            500,
        )
    try:
        # ambil data headers terlebih dahulu
        x_request, x_csrf = (
            request.headers.get("X-Requested-With"),
            request.headers.get("X-CSRF-TOKEN"),
        )
        if not x_request or not x_csrf or x_request != "XMLHttpRequest":
            return (
                jsonify({"redirect": url_for("signIn", msg="Terjadi kesalahan data")}),
                500,
            )

        status, id_status = request.json.values()
        if (
            not status
            and not id_status
            and type(status) != str
            and type(id_status) != str
        ):
            raise Exception("Terjadi Kesalahan dalam pengiriman data")

        # decode payload
        payloads = jwt.decode(cookie, secretKey, algorithms=["HS256"])
        if payloads["role"] != 1:
            return (
                jsonify(
                    {"redirect": url_for("dashboard", msg="Anda tidak memiliki akses")}
                ),
                500,
            )

        # decript id riwayat absen
        status_id = uuid_like_to_string(id_status)  

        # pemgambilan data message_id dari db faq
        message_id = db.faq.find_one(
            {"_id": ObjectId(status_id)},
            {"_id": 0, "message_id": 1, "email": 1, "no_ticket": 1, "name": 1},
        )
        # cek ditemukan / tidak
        if not message_id:
            raise Exception("Terjadi Kesalahan dalam pengambilan data")

        # sesuaikan status untuk body gmail kirim
        if status == "Diproses":
            body = f"""
            <p class='poppins-regular'>Thanks for contacts our help, Your inquiry was still in the processing stage, please wait for <strong> 2 x 24 hours </strong> for our response</p>
            """
        elif status == "Selesai":
            body = f"""
            <p class='poppins-regular'>Thanks for your waiting, Your inquiry has been resolved, Hopefully the solution we provide can help you.
            </p>
            <p class='poppins-regular'>If you have any questions, please contact us at <a class='text-info' href='{url_for("signIn",_external=True)}'>Signin</a></p>
            """
        else:
            raise Exception("Terjadi Kesalahan dalam status")

        # kirim pesan ke gmail untuk update status
        reply_message = replyGmailSender(message_id["message_id"])
        reply_message.service_gmail_api()
        reply_message.reply_message_make(
            name=message_id["name"],
            email=message_id["email"],
            no_ticket=message_id["no_ticket"],
            body=body,
        )
        response = reply_message.send_reply_message()
        # cek response yang didapatkan
        if not response:
            raise Exception("Terjadi Kesalahan dalam reply email")
        # update proses riwayatnya
        if status == "Diproses":
            result = db.faq.find_one_and_update(
                {"_id": ObjectId(status_id)},
                {
                    "$set": {
                        "status": status,
                        "message_id": {
                            "id": response["id"],
                            "threadId": response["threadId"],
                        },
                    }
                },
                {"_id": 0, "no_ticket": 1, "name": 1},
            )
        elif status == "Selesai":
            result = db.faq.find_one_and_delete(
                {"_id": ObjectId(status_id)}, {"_id": 0, "no_ticket": 1, "name": 1}
            )
        else:
            raise Exception("Terjadi Kesalahan status")
        # Cek apakah update berhasil
        if not result:
            raise Exception("Data status tidak ditemukan dan gagal di update")

        return (
            jsonify(
                {
                    "redirect": url_for(
                        "riwayat_bantuan",
                        msg=Markup(
                            f"Status bantuan dengan no <span class='poppins-semibold'>'#{result['no_ticket']}'</span> bernama <span class='poppins-semibold'>{result['name']}</span> berhasil di update"
                        ),
                        status="success",
                    )
                }
            ),
            200,
        )

    # The above code is handling exceptions related to JWT (JSON Web Token) authentication. It catches
    # different types of exceptions that can occur during JWT verification:
    except jwt.ExpiredSignatureError:
        return jsonify({"redirect": url_for("signIn", msg="Session Kadaluarsa")}), 500
    except jwt.DecodeError:
        return jsonify({"redirect": url_for("signIn", msg="Anda telah logout")}), 500
    except Exception as e:
        return jsonify({"redirect": url_for("riwayat_bantuan", msg=e.args[0])}), 500
    except IndexError as e:
        return jsonify({"redirect": url_for("riwayat_bantuan", msg=e.args[0])}), 500


@app.route("/kelola-admin", methods=["GET"])
@app.route("/kelola-admin/", methods=["GET"])
@app.route("/kelola-admin/<path1>", methods=["POST"])
@app.route("/kelola-admin/<path1>/<path2>", methods=["POST"])
def kelola_admin(path1=None, path2=None):
    """
    Fungsi ini digunakan untuk mengelola data admin dan sub admin.

    Parameter path1 dan path2 digunakan untuk mengatur operasi yang akan dilakukan terhadap data admin dan sub admin.
    Path1 dapat berisi nilai `tambah`, `edit`, `hapus`, atau `daftar`.
    Path2 digunakan untuk mengatur ID admin yang akan dihapus.

    Fungsi ini akan mengembalikan status 200 jika operasi berhasil dilakukan.
    Fungsi ini akan mengembalikan status 400 jika input tidak valid.
    Fungsi ini akan mengembalikan status 500 jika terjadi kesalahan di server.

    Contoh penggunaan:
    - Tambah data admin: /kelola-admin/tambah
    - Edit data admin: /kelola-admin/edit/<id_admin>
    - Hapus data admin: /kelola-admin/hapus/<id_admin>
    - Daftar admin: /kelola-admin/daftar
    """

    try:
        # The above Python code is checking for the presence of a CSRF token and a token key in the
        # request cookies. If the CSRF token is not found (i.e., None), it redirects the user to the
        # "signIn" route with a message indicating that the CSRF token has expired. Similarly, if the
        # token key is not found (i.e., None), it redirects the user to the "signIn" route with a
        # message indicating that the user has been logged out. This code is likely part of a web
        # application's security mechanism to ensure that valid CSRF tokens and user authentication
        # tokens are present before allowing access
        csrf_token = request.cookies.get("csrf_token")
        cookies = request.cookies.get("token_key")
        if csrf_token == None:
            return redirect(url_for("signIn", msg="csrf token Kadaluarsa"))
        if cookies == None:
            return redirect(url_for("signIn", msg="Anda Telah logout"))
        # The above Python code snippet is converting `cookies` and `csrf_token` variables from a
        # UUID-like format to a string format using the `uuid_like_to_string` function. It then checks
        # if the `csrf_token` or `cookies` are empty, and if so, it redirects the user to the "signIn"
        # route with a message indicating that either the CSRF token has expired or the cookie has
        # expired.
        cookies = uuid_like_to_string(cookies)
        csrf_token = uuid_like_to_string(csrf_token)
        if not csrf_token:
            return redirect(url_for("signIn", msg="CSRF Token Kadaluarsa"))
        if not cookies:
            return redirect(url_for("signIn", msg="Cookie Kadaluarsa"))
        # The above Python code snippet is decoding a JSON Web Token (JWT) from the `cookies` using a
        # `secretKey` with the HS256 algorithm. It then retrieves data from a MongoDB database based
        # on the decoded payload. If the role in the payload is not equal to 1 and the job is not
        # 'admin', it fetches the attendance history (`riwayat_absent`) for a specific user ID.
        # Otherwise, if the conditions are not met, it fetches the attendance history for all users.
        payloads = jwt.decode(cookies, secretKey, algorithms=["HS256"])

        # The above Python code is checking the HTTP request method. If the method is 'POST', it then
        # checks the value of the variable `path1`. If `path1` is equal to 'edit', it executes some
        # code (not shown in the snippet). If `path1` is equal to 'delete', it executes some other
        # code (not shown in the snippet). If `path1` is not equal to 'edit' or 'delete', it does
        # nothing.
        if request.method == "POST":
            # ubah data
            if path1 == "edit":
                # The above code is unpacking the values from the `request.form` object into the
                # variables `method`, `csrf_form`, `id_data_user_admin`, `nama`, `email`,
                # `departement`, and `jobs`. This allows you to access the form data submitted in a
                # Python web application using Flask or a similar framework.
                (
                    method,
                    csrf_form,
                    id_data_user_admin,
                    nama,
                    email,
                    departement,
                    jobs,
                ) = request.form.values()

                # The above Python code snippet is performing input validation and conditional checks
                # based on certain conditions. Here is a breakdown of what the code is doing:
                if method != "PUT":
                    return redirect(
                        url_for("kelola_admin", msg="This method form not allowed")
                    )
                if csrf_form == None or csrf_form == "":
                    return redirect(
                        url_for("kelola_admin", msg="Your request not in page")
                    )
                if id_data_user_admin == None or id_data_user_admin == "":
                    raise Exception("This data is invalid")
                if nama == None or nama == "":
                    raise ValueError("Your name cannot be empty")
                if email == None or nama == "":
                    raise ValueError("Your email cannot be empty")
                if departement == None or departement == "":
                    raise ValueError("Your departement cannot be selected")
                if jobs == None or jobs == "":
                    raise ValueError("Your jobs cannot be selected")
                if payloads["role"] != 1 and payloads["jobs"] not in ("Admin"):
                    raise Exception("Anda tidak memiliki hak akses")

                # cari departement
                departement_payloads = db.users.find_one(
                    {"_id": ObjectId(payloads["_id"])}, {"_id": 0, "departement": 1}
                )
                # jika bukan departemen superuser
                if departement_payloads["departement"] not in ("Superuser"):
                    raise Exception("Anda tidak memiliki hak akses")

                # decript id riwayat absen
                user_admin_id = uuid_like_to_string(id_data_user_admin)  

                # cek ada / tidak
                if not user_admin_id:
                    raise ValueError("Terjadi kesalahan dalam encode decrypt")

                # update database
                result = db.users.find_one_and_update(
                    {"_id": ObjectId(user_admin_id)},
                    {
                        "$set": {
                            "nama": nama,
                            "email": email,
                            "jobs": jobs,
                            "departement": departement,
                        }
                    },
                    {"_id": 0, "nama": 1},
                )

                # lakukan pengecekan dalam perubahan data
                if not result:
                    raise ValueError("Data atas nama " + nama + " failed patch")
                return redirect(
                    url_for(
                        "kelola_admin",
                        msg="Data atas nama " + result["nama"] + " success edit",
                        result="success",
                    )
                )

            # delete data
            elif path1 == "delete":
                # jika path2 tidak diinisialisasi
                if path2 == None:
                    return redirect(url_for("notFound"))

                # cari departement
                departement_payloads = db.users.find_one(
                    {"_id": ObjectId(payloads["_id"])}, {"_id": 0, "departement": 1}
                )

                # cek role=1 dan jobs = admin
                if (
                    payloads["role"] != 1
                    and payloads["jobs"] not in ("Admin")
                    and departement_payloads["departement"] not in ("Superuser")
                ):
                    raise ("Anda tidak memiliki hak akses")

                # ambil form data
                method, csrf_form = request.form.values()

                # The code snippet provided is written in Python and it checks two conditions.
                if method != "DELETE":
                    return redirect(url_for("notFound"))
                if csrf_form == None or csrf_form == "":
                    return redirect(url_for("signIn", msg="Your request not in page"))

                # decript id user admin id
                user_admin_id_delete = uuid_like_to_string(path2)

                # cek ada / tidak
                if not user_admin_id_delete:
                    raise ValueError("Terjadi kesalahan dalam encode decrypt")

                # delete data
                result = db.users.find_one_and_delete(
                    {"_id": ObjectId(user_admin_id_delete)}
                )

                # lakukan pengecekan dalam perubahan data
                if not result:
                    raise Exception(
                        "Data atas nama " + result["nama"] + " failed delete"
                    )
                return redirect(
                    url_for(
                        "kelola_admin",
                        msg="Data atas nama " + result["nama"] + " success delete",
                        result="success",
                    )
                )

            # selain itu
            else:
                raise Exception("this path url doesnt exist")

        # The above code is attempting to retrieve a value from the query parameters of a request in a
        # web application. It is trying to get the value associated with the key 'msg' from the
        # request arguments.
        msg = request.args.get("msg")
        result = request.args.get("result")

        # ambil data untuk navbar
        data = db.users.find_one(
            {"_id": ObjectId(payloads["_id"])},
            {
                "_id": 0,
                "jobs": 1,
                "role": 1,
                "nama": 1,
                "photo_profile": 1,
                "departement": 1,
            },
        )
        if not data:
            raise Exception("Terjadi kesalahan data")

        # cek apakah dia bukan tipe admin
        if (
            payloads["role"] != 1
            or data["departement"] not in ("Superuser")
        ):
            return redirect(url_for("notFound"))
        
        query = {
            "jobs": {"$in": ["Admin", "Sub Admin"]},
            "role": 1,
            "_id": {"$ne": ObjectId(payloads["_id"])}
        }
        # Ambil informasi user login
        is_superuser = data.get("departement") == "Superuser"
        is_sub_admin = data.get("jobs") == "Sub Admin"
        
        # Superuser Sub Admin hanya bisa lihat Sub Admin dari Superuser
        if is_superuser and is_sub_admin:
            query["$or"] = [
                {"departement": {"$ne": "Superuser"}},  # boleh dari luar Superuser
                {"departement": "Superuser", "jobs": "Sub Admin"}  # hanya Sub Admin Superuser
            ]

        # cari data admin
        data_admin_all = list(
            db.users.find(
                query,
                {"password": 0, "role": 0},
            )
        )

        # cek data admin_all
        if data_admin_all:
            for i in data_admin_all:
                i["_id"] = string_to_uuid_like(
                    str(i["_id"])
                )
                if not i["_id"]:
                    raise Exception("Terjadi Kesalahan Data dalam encrypt")

        # render template get
        return render_template(
            "kelolaAdmin.html",
            data=data,
            data_user_admin=data_admin_all,
            msg=msg,
            result=result,
        )

    # handle error
    except jwt.ExpiredSignatureError:
        return redirect(url_for("signIn", msg="Session Kadaluarsa"))
    except jwt.DecodeError:
        return redirect(url_for("signIn", msg="Anda telah logout"))
    except ValueError as e:
        return redirect(
            url_for("kelola_admin", msg=e.args[0] if e.args else "value error occurred")
        )
    except Exception as e:
        return redirect(
            url_for(
                "dashboard", msg=e.args[0] if e.args else "An unexpected error occurred"
            )
        )
    except IndexError as e:
        return redirect(
            url_for("kelola_admin", msg=e.args[0] if e.args else "Index error occurred")
        )


# kelola admin export
@app.route("/kelola-admin/<path1>", methods=["GET"])
@app.route("/kelola-admin/<path1>/", methods=["GET"])
def kelola_admin_export(path1):
    """
    Kelola Admin Export
    ---
    tags:
      - Export
    summary: Kelola Admin Export
    description: This endpoint is used to export admin data in different formats. It requires valid authentication cookies and a CSRF token to access the data.
    parameters:
      - name: path1
        in: path
        required: true
        enum:
          - excel
          - pdf
        description: The format in which to export the data. Accepted values are "excel" or "pdf".
        type: string
        example: excel
    responses:
      200:
        description: The data export is successful.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "Data exported successfully"
            status:
              type: string
              example: "success"
      401:
        description: CSRF token or cookie is invalid or expired.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "CSRF token Kadaluarsa"
            status:
              type: string
              example: "error"
      404:
        description: The specified export path is invalid.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "Export path not found"
            status:
              type: string
              example: "error"
      500:
        description: An error occurred on the server.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "An error occurred while exporting data"
            status:
              type: string
              example: "error"
    """
    if path1 not in ["excel", "pdf"]:
        return redirect(url_for("notFound"))
    try:
        # ambil cookie dan payload
        cookie = request.cookies.get("token_key")
        csrf_token = request.cookies.get("csrf_token")
        if csrf_token == None:
            return redirect(url_for("signIn", msg="csrf token Kadaluarsa"))
        if cookie == None or cookie == "":
            return redirect(url_for("signIn", msg="Anda Telah logout"))
        cookie = uuid_like_to_string(cookie)
        if not cookie:
            return redirect(url_for("signIn", msg="Cookie Kadaluarsa"))
        csrf_token = uuid_like_to_string(csrf_token)
        if not csrf_token:
            return redirect(url_for("signIn", msg="CSRF Token Kadaluarsa"))
        # decode payload
        payloads = jwt.decode(cookie, secretKey, algorithms=["HS256"])

        # The above Python code is checking if certain conditions are met before raising an exception
        # with the message "Anda tidak punya akses" which translates to "You do not have access" in
        # English. The conditions being checked are:
        # 1. The value of the "role" key in the payloads dictionary is not equal to 1.
        # 2. The value of the "jobs" key in the payloads dictionary is not equal to "admin".
        # 3. The value of the "departement" key in the payloads dictionary is not in the tuple
        # ("Superuser", "Subuser").

        data_user = db.users.find_one(
            {"_id": ObjectId(payloads["_id"])},
            {
                "_id": 0,
                "jobs": 1,
                "role": 1,
                "nama": 1,
                "photo_profile": 1,
                "departement": 1,
            },
        )

        if (
            payloads["role"] != 1
            and payloads["jobs"] != "admin"
            and data_user["departement"] != "Superuser"
        ):
            raise Exception("Anda tidak punya akses")

        # cek database
        result = list(
            db.users.find(
                {"jobs": {"$in": ["Admin", "Sub Admin"]}},
                {"_id": 0, "password": 0, "role": 0},
            )
        )
        if result == None or result == "" or not result:
            return redirect(url_for("kelola_admin", msg="Data user not found"))
        # save excel data
        file_path = os.path.join(app.root_path, "static", "doc")
        wb = load_workbook(file_path + "/excel/template_kelola_admin.xlsx")
        ws = wb.active
        column_widths, start, stop = convert_to_excel(
            ws, result=result, currentPage="Kelola Admin"
        )

        # Cek apakah path dapat ditulis
        if not os.access(file_path, os.W_OK):
            # Jika tidak bisa menulis ke folder static, simpan ke /tmp
            print("Tidak dapat menulis ke folder static. Menyimpan ke /tmp...")
            file_path = "/tmp"
            os.makedirs(os.path.join(file_path, "excel"), exist_ok=True)
            os.makedirs(os.path.join(file_path, "pdf"), exist_ok=True)
            file_path = os.path.join(file_path)
        # save for extractor excel
        wb.save(file_path + "/excel/data_admin_sub_admin.xlsx")
        # excel
        if path1 == "excel":
            # Save the modified workbook to BytesIO
            output = BytesIO()
            wb.save(output)
            # Mengirim file Excel sebagai response download
            output.seek(0)
            return send_file(
                output,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="data_admin_sub_admin_winnicode.xlsx",
            )

        # pdf
        elif path1 == "pdf":
            # The above code is written in Python and seems to be using a library called "PDF" to
            # create a PDF document. It specifies the page size as A4 and then adds a new page to the
            # PDF. Finally, it calls the `create_pdf` method with parameters `start - 1`, `stop`,
            # `column_widths`, and `ws` to generate content in the PDF document.
            pdf = PDF("portrait", "mm", "A4")
            pdf.add_page()
            pdf.create_pdf(
                start - 1, stop, column_widths, ws, currentPage="Kelola Admin"
            )
            # Buat objek BytesIO
            outputs = BytesIO()
            pdf.output(file_path + "/pdf/data_admin_sub_admin.pdf")
            # Kirim file PDF sebagai attachment
            return send_file(
                file_path + "/pdf/data_admin_sub_admin.pdf",
                mimetype="application/pdf",
                as_attachment=True,
                download_name="data_admin_sub_admin_winnicode.pdf",
            )

        # jika tidak ada
        else:
            return redirect(url_for("notFound"))

    # handle error
    except jwt.ExpiredSignatureError:
        return redirect(url_for("signIn", msg="Session Kadaluarsa"))
    except jwt.DecodeError:
        return redirect(url_for("signIn", msg="Anda telah logout"))
    except Exception as e:
        return redirect(url_for("dashboard", msg=e.args[0]))


# add account admin / sub admin
@app.route("/kelola-admin/create-account", methods=["POST"])
def dashboardAdminCreateAccount():
    """
    Buat Akun admin / sub admin
    ---
    tags:
      - Admin / Sub Admin
    summary: Create Admin Account
    description: This endpoint is used to create a new admin or sub-admin account. It requires valid authentication cookies and a CSRF token to perform the operation.
    parameters:
      - name: csrf_token
        in: formData
        required: true
        description: CSRF token from the form.
        type: string
      - name: nama
        in: formData
        required: true
        description: Full name of the user.
        type: string
      - name: email
        in: formData
        required: true
        description: Email address of the user.
        type: string
      - name: departement
        in: formData
        required: true
        enum:
          - Superuser
          - Mentor
        description: Department of the user.
        type: string
      - name: jobs
        in: formData
        required: true
        enum:
          - Admin
          - Sub Admin
        description: Job title of the user (e.g., "Admin", "Sub Admin").
        type: string
      - name: password
        in: formData
        required: true
        description: Password for the new account.
        type: string
      - name: password2
        in: formData
        required: true
        description: Password confirmation for validation.
        type: string
    responses:
      200:
        description: The account creation is successful.
      400:
        description: The request was invalid or cannot be served.
      401:
        description: CSRF token or cookie is invalid or expired.
      403:
        description: The user does not have access to create an account.
      500:
        description: An error occurred on the server.
    """

    try:
        # The above Python code is checking for the presence of a CSRF token and a token key in the
        # request cookies. If the CSRF token is not found (i.e., None), it redirects the user to the
        # "signIn" route with a message indicating that the CSRF token has expired. Similarly, if the
        # token key is not found (i.e., None), it redirects the user to the "signIn" route with a
        # message indicating that the user has been logged out. This code is likely part of a web
        # application's security mechanism to ensure that valid CSRF tokens and user authentication
        # tokens are present before allowing access
        csrf_token = request.cookies.get("csrf_token")
        cookies = request.cookies.get("token_key")
        if csrf_token == None:
            return redirect(url_for("signIn", msg="csrf token Kadaluarsa"))
        if cookies == None:
            return redirect(url_for("signIn", msg="Anda Telah logout"))
        # The above Python code snippet is converting `cookies` and `csrf_token` variables from a
        # UUID-like format to a string format using the `uuid_like_to_string` function. It then checks
        # if the `csrf_token` or `cookies` are empty, and if so, it redirects the user to the "signIn"
        # route with a message indicating that either the CSRF token has expired or the cookie has
        # expired.
        cookies = uuid_like_to_string(cookies)
        csrf_token = uuid_like_to_string(csrf_token)
        if not csrf_token:
            return redirect(url_for("signIn", msg="CSRF Token Kadaluarsa"))
        if not cookies:
            return redirect(url_for("signIn", msg="Cookie Kadaluarsa"))
        # The above Python code snippet is decoding a JSON Web Token (JWT) from the `cookies` using a
        # `secretKey` with the HS256 algorithm. It then retrieves data from a MongoDB database based
        # on the decoded payload. If the role in the payload is not equal to 1 and the job is not
        # 'admin', it fetches the attendance history (`riwayat_absent`) for a specific user ID.
        # Otherwise, if the conditions are not met, it fetches the attendance history for all users.
        payload = jwt.decode(cookies, secretKey, algorithms=["HS256"])
        # cek payload
        if not payload:
            raise Exception("Invalid token")

        data_user = db.users.find_one(
            {"_id": ObjectId(payload["_id"])},
            {"_id": 0, "departement": 1, "jobs": 1, "role": 1},
        )

        # cek hak akses selama cookie
        if (
            payload["role"] != 1
            and payload["jobs"] not in ("Admin")
            and data_user["departement"] not in ("Superuser")
        ):
            raise Exception("Anda tidak memiliki akses")

        # inisiasi validasi inputan
        csrf_token, nama, email, departement, jobs, password, password2 = (
            request.form.values()
        )

        # validasi terlebih dahulu kosong atau tidak
        if csrf_token.strip() == "" or csrf_token == None:
            raise Exception("Your request not in page")
        if nama.strip() == "" or nama == None:
            raise ValueError("your name is empty, please insert in")
        if email.strip() == "" or email == None:
            raise ValueError("your email is empty, please insert in")
        if departement.strip() == "" or departement == None:
            raise ValueError("your departement is empty, please insert in")
        if jobs.strip() == "" or jobs == None:
            raise ValueError("your jobs is empty, please insert in")
        if password.strip() == "" or password == None:
            raise ValueError("your password is empty, please insert in")
        if password2.strip() == "" or password2 == None:
            raise ValueError("your confirm password is empty, please insert in")

        # jika tidak lanjut kesini
        # cek inputan password dengan konfirm password sama atau tidak
        if password.strip() != password2.strip():
            raise ValueError("your password and confirm password not same")

        # cek email apakah sudah ada di database
        if db.users.find_one(
            {
                "email": email.strip(),
                "nama": nama.strip(),
                "jobs": jobs.strip(),
                "departement": departement.strip(),
            }
        ):
            raise ValueError("email or name already exist")

        # ambil callback insert user
        result = db.users.insert_one(
            {
                "nama": nama.strip(),
                "email": email.strip().lower(),
                "password": hashlib.sha256(password.strip().encode()).hexdigest(),
                "departement": departement.strip(),
                "jobs": jobs.strip(),
                "role": 1,
                "photo_profile": "https://i.ibb.co.com/5Yd94zx/user.png",
            }
        )
        # berhasi/ tidak
        if not result:
            raise ValueError("Data tidak berhasil disimpan")
        # render
        return redirect(
            url_for(
                "kelola_admin",
                msg="Data atas nama " + str(nama.strip()) + " Berhasil Disimpan",
                result="success",
            )
        )
    # handling error
    except jwt.ExpiredSignatureError:
        return redirect(url_for("signIn", msg="Session Kadaluarsa"))
    except jwt.DecodeError:
        return redirect(url_for("signIn", msg="Anda telah logout"))
    except ValueError as e:
        return redirect(url_for("kelola_admin", msg=e.args[0]))
    except Exception as e:
        return redirect(url_for("dashboard", msg=e.args[0]))


# mengedit data karyawan melalui admin
@app.route("/dashboard/admin/edit", methods=["POST"])
def dashboardAdminEdit():
    """
    Edit Data Karyawan melalui Admin
    ---
    tags:
      - Admin / Sub Admin
    summary: Edit Karyawan Data
    description: Endpoint ini digunakan untuk melakukan edit data karyawan / magang melalui admin. ini membutuhkan autentikasi valid cookie dan csrf token.
    parameters:
      - name: _method
        in: formData
        required: true
        enum:
          - PUT
        description: Should be "PUT".
        type: string
      - name: csrf_token
        in: formData
        required: true
        description: CSRF token that is valid.
        type: string
      - name: nik
        in: formData
        required: true
        description: Nik of the employee to edit.
        type: string
      - name: nama
        in: formData
        required: true
        description: Name of the employee to edit.
        type: string
      - name: email
        in: formData
        required: true
        description: Email of the employee to edit.
        type: string
      - name: departement
        in: formData
        required: true
        enum:
          - Superuser
          - Mentor
        description: Department of the employee to edit.
        type: string
      - name: jobs
        in: formData
        required: true
        enum:
          - Admin
          - Sub Admin
        description: Job title of the employee to edit.
        type: string
      - name: start_date
        in: formData
        required: true
        description: "Start date of the employee (format: YYYY-MM-DD)."
        type: string
      - name: end_date
        in: formData
        required: true
        description: "End date of the employee (format: YYYY-MM-DD)."
        type: string
      - name: start_time
        in: formData
        required: true
        description: "Start time of the employee (format: hh:mm:ss)."
        type: string
      - name: end_time
        in: formData
        required: true
        description: "End time of the employee (format: hh:mm:ss)."
        type: string
    responses:
      200:
        description: The request is successful.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "Employee data updated successfully"
            status:
              type: string
              example: "success"
      400:
        description: Input is invalid.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "Invalid input data"
            status:
              type: string
              example: "error"
      401:
        description: CSRF token or cookie is invalid or expired.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "CSRF token Kadaluarsa"
            status:
              type: string
              example: "error"
      500:
        description: An error occurred on the server.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "An error occurred while updating employee data"
            status:
              type: string
              example: "error"
    """

    # apakah method put /nukan
    if request.form.get("_method") != "PUT":
        return redirect(url_for("dashboard"))

    # cek csrf token
    csrf_token = request.form.get("csrf_token")
    # cek cookie token
    cookies = request.cookies.get("token_key")

    # ambil data
    nama = request.form.get("nama")
    email = request.form.get("email")
    nik = request.form.get('nik')
    departement = request.form.get("departement")
    jobs = request.form.get("jobs")
    start_date = request.form.get("start_date")
    end_date = request.form.get("end_date")
    start_time = request.form.get("start_time")
    end_time = request.form.get("end_time")

    # cek kosong / tidak
    if csrf_token == "" or csrf_token == None:
        raise ValueError("csrf token Kadaluarsa")
    try:
        # cek string kosong atau tidak
        if (
            nama == ""
            or email == ""
            or nik == ""
            or departement == "None"
            or jobs == "None"
            or start_date == ""
            or end_date == ""
            or start_time == ""
            or end_time == ""
        ):
            return redirect(
                url_for("dashboard", msg="data tidak boleh ada yang kosong")
            )
        # decode uuid
        cookies = uuid_like_to_string(cookies)
        
        nik = int(nik)
        if type(nik) != int:
            return redirect(url_for("dashboard", msg="NIK harus angka"))
        
        # cek berhasul / tidak
        if not cookies:
            raise ValueError("CSRF decode error")
        # decode jwt
        payloads = jwt.decode(cookies, secretKey, algorithms=["HS256"])

        data_user = db.users.find_one(
            {"_id": ObjectId(payloads["_id"])},
            {"_id": 0, "departement": 1, "jobs": 1, "role": 1},
        )

        if not data_user:
            raise Exception("Data user akses tidak ditemukan")

        # cek hak akses selama cookie
        if (
            payloads["role"] != 1
            and payloads["jobs"] not in ("Admin","Sub Admin")
            and data_user["departement"] not in ("Superuser")
        ):
            raise Exception("Anda tidak memiliki akses")

        # lakukan update
        result = db.users.update_one(
            {"email": email, "nama": nama},
            {
                "$set": {
                    "nik": nik,
                    "departement": departement,
                    "mulai_kerja": datetime.datetime.strptime(start_date, "%Y-%m-%d")
                    .strftime("%d %B %Y")
                    .lower(),
                    "akhir_kerja": datetime.datetime.strptime(end_date, "%Y-%m-%d")
                    .strftime("%d %B %Y")
                    .lower(),
                    "waktu_awal_kerja": datetime.datetime.strptime(start_time, "%H:%M")
                    .strftime("%H.%M")
                    .lower(),
                    "waktu_akhir_kerja": datetime.datetime.strptime(end_time, "%H:%M")
                    .strftime("%H.%M")
                    .lower(),
                    "jobs": jobs,
                    "work_hours": (
                        datetime.datetime.strptime(end_time, "%H:%M")
                        - datetime.datetime.strptime(start_time, "%H:%M")
                    ).seconds
                    // 3600,
                }
            },
        )
        # berhasil atau tidak dalam melakukan update user
        if result.matched_count > 0:
            return redirect(
                url_for(
                    "dashboard",
                    msg="data karyawan / magang berhasil di edit",
                    result="success",
                )
            )
        return redirect(
            url_for("dashboard", msg="data karyawan / magang gagal di edit")
        )
    # handle error
    except jwt.ExpiredSignatureError:
        return redirect(url_for("signIn", msg="Session Kadaluarsa"))
    except jwt.DecodeError:
        return redirect(url_for("signIn", msg="Anda telah logout"))
    except ValueError as e:
        return redirect(url_for("signIn", msg=e.args[0]))
    except Exception as e:
        return redirect(url_for("dashboard", msg=e.args[0]))


# delete user karyawan / magang melalui admin
@app.route("/dashboard/admin/delete/<id>", methods=["POST"])
def adminDelete(id):
    """
    Delete account Karyawan / Magang
    ---
    tags:
      - Admin / Sub Admin
    summary: Delete User Karyawan / Magang
    description: Endpoint ini digunakan untuk menghapus akun karyawan / magang melalui admin. ini membutuhkan autentikasi valid cookie dan csrf token.
    parameters:
      - name: id
        in: path
        required: true
        description: The ID of the user to be deleted.
        type: string
    responses:
      200:
        description: The request is successful.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "Data karyawan / magang berhasil dihapus"
            status:
              type: string
              example: "success"
      400:
        description: Input is invalid.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "Invalid input data"
            status:
              type: string
              example: "error"
      401:
        description: CSRF token or cookie is invalid or expired.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "CSRF token Kadaluarsa"
            status:
              type: string
              example: "error"
      500:
        description: An error occurred on the server.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "An error occurred while deleting the user"
            status:
              type: string
              example: "error"
    """

    # The above code is checking if the value of the "_method" key in the request form is not equal to
    # "DELETE". If the condition is true, the code inside the if block will be executed.
    if request.form.get("_method") != "DELETE":
        return redirect(url_for("dashboard"))
    # cek cookie token
    cookies = request.cookies.get("token_key")
    try:
        # The above Python code is checking if the value of the "csrf_token" parameter in the form
        # data is an empty string or None. If the value is empty or None, it raises a ValueError with
        # the message "csrf token Kadaluarsa". This is a common security measure to ensure that a valid
        # CSRF token is present in the form data to prevent CSRF attacks.
        if (
            request.form.get("csrf_token") == ""
            or request.form.get("csrf_token") == None
        ):
            raise ValueError("csrf token Kadaluarsa")
        # decode uuid
        cookies = uuid_like_to_string(cookies)

        # The code is checking if the variable `cookies` is falsy. If `cookies` is falsy, the
        # condition will evaluate to `True`, otherwise it will evaluate to `False`.
        if not cookies:
            raise ValueError("decode error")

        # The above code is using the `jwt.decode` function to decode a JSON Web Token (JWT) stored in
        # the `cookies` variable using the specified `secretKey` and algorithm "HS256". This function
        # will verify the signature of the JWT and return the decoded payload if the signature is
        # valid.
        payloads = jwt.decode(cookies, secretKey, algorithms=["HS256"])

        data_user = db.users.find_one(
            {"_id": ObjectId(payloads["_id"])},
            {"_id": 0, "departement": 1, "jobs": 1, "role": 1},
        )
        if not data_user:
            raise Exception("Data anda tidak ditemukan")

        # cek hak akses selama cookie
        if (
            payloads["role"] != 1
            and payloads["jobs"] not in ("Admin")
            and data_user["departement"] in ("Superuser")
        ):
            raise Exception("Anda tidak memiliki akses")

        # The above code is written in Python and it is performing two operations using MongoDB
        # database:
        result1 = db.users.find_one_and_delete({"_id": ObjectId(id)})
        result2 = db.absen_magang.delete_many({"user_id": ObjectId(id)})
        result3 = db.tasks.delete_many({"user_id": ObjectId(id)})
        result4 = db.angka_notif.delete_many({result1['email']})
        
        if not result1:
            raise Exception("Data user yang akan dihapus tidak ditemukan")

        # The above Python code is checking if the `deleted_count` attribute of `result1` is greater
        # than 0 and the `deleted_count` attribute of `result2` is greater than or equal to 0. If both
        # conditions are met, it will redirect the user to the "dashboard" route with a success
        # message indicating that employee/intern data has been successfully deleted.
        if (
            result1.deleted_count > 0
            and result2.deleted_count >= 0
            and result3.deleted_count >= 0
            and result4.deleted_count >= 0
        ):
            return redirect(
                url_for(
                    "dashboard",
                    msg="data karyawan / magang berhasil di hapus",
                    result="success",
                )
            )

        # The above code is a Python code snippet that is using the Flask framework. It is performing
        # a redirect to the "dashboard" route with a message parameter "msg" set to "data karyawan /
        # magang gagal di hapus". This means that it is redirecting the user to the dashboard route
        # with a message indicating that the deletion of employee/intern data has failed.
        return redirect(
            url_for("dashboard", msg="data karyawan / magang gagal di hapus")
        )
    # The above code is handling different exceptions that may occur in a Python application.
    except jwt.DecodeError:
        return redirect(url_for("signIn", msg="Anda telah logout"))
    except jwt.ExpiredSignatureError:
        return redirect(url_for("signIn", msg="Session Kadaluarsa"))
    except ValueError as e:
        return redirect(url_for("signIn", msg=e.args[0]))
    except Exception as e:
        return redirect(url_for("dashboard", msg=e.args[0]))


@app.route("/dashboard/admin/<path>", methods=["GET"])
def export(path):
    """
    Export data karyawan / magang ke excel atau pdf
    ---
    tags:
      - Export
    summary: Ekspor Data Admin
    description: Endpoint ini digunakan untuk mengekspor data karyawan / magang dalam berbagai format. Memerlukan cookie autentikasi dan token CSRF yang valid untuk mengakses data.
    parameters:
      - name: path
        in: path
        enum:
          - pdf
          - excel
        required: true
        description: Format untuk mengekspor data. Nilai yang diterima adalah "excel" atau "pdf".
        type: string
    responses:
      200:
        description: Ekspor data berhasil.
      401:
        description: Token CSRF atau cookie tidak valid atau telah kedaluwarsa.
      404:
        description: Jalur ekspor yang ditentukan tidak valid.
      500:
        description: Terjadi kesalahan di server.
    """

    if path not in ["excel", "pdf"]:
        return redirect(url_for("notFound"))
    try:
        # ambil cookie dan payload
        cookie = request.cookies.get("token_key")
        csrf_token = request.cookies.get("csrf_token")
        if csrf_token == None:
            return redirect(url_for("signIn", msg="csrf token Kadaluarsa"))
        if cookie == None or cookie == "":
            return redirect(url_for("signIn", msg="Anda Telah logout"))
        cookie = uuid_like_to_string(cookie)
        if not cookie:
            return redirect(url_for("signIn", msg="Cookie Kadaluarsa"))
        csrf_token = uuid_like_to_string(csrf_token)
        if not csrf_token:
            return redirect(url_for("signIn", msg="CSRF Token Kadaluarsa"))
        # decode payload
        payloads = jwt.decode(cookie, secretKey, algorithms=["HS256"])

        # The above Python code is checking if certain conditions are met before raising an exception
        # with the message "Anda tidak punya akses" which translates to "You do not have access" in
        # English. The conditions being checked are:
        # 1. The value of the "role" key in the payloads dictionary is not equal to 1.
        # 2. The value of the "jobs" key in the payloads dictionary is not equal to "admin".
        # 3. The value of the "departement" key in the payloads dictionary is not in the tuple
        # ("Superuser", "Subuser").
        if (
            payloads["role"] != 1
            and payloads["jobs"] != "admin"
            and payloads["departement"] not in ("Superuser", "Sub user")
        ):
            raise Exception("Anda tidak punya akses")
        # cek database
        result = list(
            db.users.find(
                {"jobs": {"$in": ["Karyawan", "Magang"]}},
                {"_id": 0, "password": 0, "role": 0, "work_hours": 0},
            )
        )
        if result == None or result == "" or not result:
            raise Exception("Data user not found")
        # save excel data
        file_path = os.path.join(app.root_path, "static", "doc")
        wb = load_workbook(file_path + "/excel/template_data_karyawan.xlsx")
        ws = wb.active
        column_widths, start, stop = convert_to_excel(ws, result=result)

        # Cek apakah path dapat ditulis
        if not os.access(file_path, os.W_OK):
            # Jika tidak bisa menulis ke folder static, simpan ke /tmp
            file_path = "/tmp"
            os.makedirs(os.path.join(file_path, "excel"), exist_ok=True)
            os.makedirs(os.path.join(file_path, "pdf"), exist_ok=True)
            file_path = os.path.join(file_path)
            print("Tidak dapat menulis ke folder static. Menyimpan ke /tmp...")

        # save for extractor excel
        wb.save(file_path + "/excel/data_karyawan.xlsx")
        # excel
        if path == "excel":
            # Save the modified workbook to BytesIO
            output = BytesIO()
            wb.save(output)
            # Mengirim file Excel sebagai response download
            output.seek(0)
            return send_file(
                output,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="data_karyawan_winnicode.xlsx",
            )
        # pdf
        elif path == "pdf":
            # The above code is written in Python and seems to be using a library called "PDF" to
            # create a PDF document. It specifies the page size as A4 and then adds a new page to the
            # PDF. Finally, it calls the `create_pdf` method with parameters `start - 1`, `stop`,
            # `column_widths`, and `ws` to generate content in the PDF document.
            pdf = PDF("L", "mm", "A4")
            pdf.add_page()
            pdf.create_pdf(start - 1, stop, column_widths, ws)
            # Buat objek BytesIO
            output = BytesIO()
            pdf.output(file_path + "/pdf/data_karyawan.pdf")
            # Kirim file PDF sebagai attachment
            return send_file(
                file_path + "/pdf/data_karyawan.pdf",
                mimetype="application/pdf",
                as_attachment=True,
                download_name="data_karyawan_winnicode.pdf",
            )
        else:
            return redirect(url_for("notFound")), 404
    # The above code is handling exceptions related to JWT (JSON Web Token) authentication. It catches
    # different types of exceptions that can occur during JWT verification:
    except jwt.ExpiredSignatureError:
        return redirect(url_for("signIn", msg="Session Kadaluarsa"))
    except jwt.DecodeError:
        return redirect(url_for("signIn", msg="Anda telah logout"))
    except Exception as e:
        return redirect(url_for("dashboard", msg=e.args[0]))


# task page
@app.route("/task", methods=["GET"])
def task():
    """
    Task Page
    ---
    tags:
      - Magang / Karyawan
      - Admin / Sub Admin
    summary: Halaman Tugas
    description: Endpoint ini menampilkan halaman manajemen tugas untuk pengguna dan admin. Memerlukan cookie autentikasi dan token CSRF yang valid.
    parameters:
      - name: status
        in: query
        required: false
        description: Status dari tugas.
        type: string
      - name: msg
        in: query
        required: false
        description: Pesan terkait tugas.
        type: string
      - name: path1
        in: query
        required: false
        description: Jalur untuk mengarahkan halaman ['Karyawan', 'Magang', id_custom, ''].
        type: string
    responses:
      200:
        description: Halaman tugas berhasil ditampilkan.
        schema:
          type: object
          properties:
            data:
              type: object
              description: Data pengguna dan informasi tugas.
      401:
        description: Akses tidak sah karena token CSRF atau cookie tidak valid atau telah kedaluwarsa.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "Token CSRF kedaluwarsa"
            status:
              type: string
              example: "error"
      500:
        description: Terjadi kesalahan di server.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "Terjadi kesalahan yang tidak terduga"
            status:
              type: string
              example: "error"
    """

    # The above Python code snippet is checking for the presence and validity of a CSRF token and a
    # cookie in a web request. Here is a breakdown of the code:
    cookie = request.cookies.get("token_key")
    csrf_token = request.cookies.get("csrf_token")
    if csrf_token == None:
        return redirect(url_for("signIn", msg="csrf token Kadaluarsa"))
    if cookie == None:
        return redirect(url_for("signIn", msg="Anda Telah logout"))

    # decode uuid
    cookie = uuid_like_to_string(cookie)
    csrf_token = uuid_like_to_string(csrf_token)
    if not csrf_token:
        return redirect(url_for("signIn", msg="CSRF Token Kadaluarsa"))
    if not cookie:
        return redirect(url_for("signIn", msg="Cookie Kadaluarsa"))

    # check status dan msg
    result = request.args.get("status")
    msg = request.args.get("msg")

    try:
        path1 = request.args.get("path1")
        table_heading = []
        data_user_all = []

        # decode payload
        payload = jwt.decode(cookie, secretKey, algorithms=["HS256"])
        # ambil data dari db
        data = db.users.find_one(
            {
                "_id": ObjectId(payload["_id"]),
                "jobs": payload["jobs"],
                "role": payload["role"],
            }
        )
        # jika data kosong
        if not data:
            return redirect(url_for("signIn", msg="Anda bukan bagian dari WinniCode"))

        # admin /sub admin
        if payload["role"] == 1 and payload["jobs"] in ("Admin", "Sub Admin"):
            # user selain payload
            data_user_for_add = list(
                db.users.find(
                    {
                        "_id": {"$ne": ObjectId(payload["_id"])},
                        "role": {"$ne": 1},
                    },  # bukan id payload dan rolwnya bukan angka 1
                    {
                        "_id": 0,
                        "nama": 1,
                        "email": 1,
                        "jobs": 1,
                        "departement": 1,
                    },  # ambil nama , email, jobs dan depratement
                )
            )

            # ambil data ynag ditampilkan
            # path1 kosong
            if not path1:
                data_user_all = list(
                    db.users.aggregate(
                        [
                            {"$match": {"jobs": {"$nin": ["Admin", "Sub Admin"]}}},
                            {
                                "$group": {
                                    "_id": "$jobs",
                                    "document": {"$first": "$$ROOT"},
                                }
                            },
                            {"$replaceRoot": {"newRoot": "$document"}},
                            {"$project": {"_id": 1, "jobs": 1}},
                        ]
                    )
                )
                
            # path 1 karyawan / magang
            elif path1 in ("Karyawan", "Magang"):
                data_user_all = list(
                    db.users.find(
                        {"jobs": path1},
                        {"_id": 1, "nama": 1, "email": 1, "departement": 1},
                    )
                )

            # selain itu
            else:
                data_user_all = list(
                    db.tasks.find(
                        {"user_id": ObjectId(uuid_like_to_string(str(path1)))},
                        {"user_id": 0},
                    )
                )

        # karyawan / magang
        elif payload["role"] == 3 and payload["jobs"] in ("Karyawan", "Magang"):
            data_user_all = list(
                db.tasks.find(
                    {"user_id": ObjectId(payload["_id"])},
                    {"accepted": 0, "user_id": 0},
                )
            )

        # iterasiins
        for user in data_user_all:
            user["_id"] = string_to_uuid_like(str(user["_id"]))
            table_heading = [x for x in user.keys() if x != "_id"]
            if payload["role"] == 3 and payload["jobs"] in ("Karyawan", "Magang"):
                if get_time_zone_now() >= user["deadline"]:
                    user["dataSort"] = "true"

        # tampilkan
        return render_template(
            "task.html",
            data=data,
            enumerate=enumerate,
            list=list,
            type=type,
            datetime=datetime.datetime,
            path1=path1,
            data_user_all=data_user_all,
            table_heading=table_heading,
            result=result,
            data_user_for_add=data_user_for_add if payload["role"] == 1 else [],
            msg=msg,
        )
    # The above code is handling exceptions related to JWT (JSON Web Token) in a Python application.
    # Specifically, it is catching `jwt.ExpiredSignatureError` and `jwt.DecodeError` exceptions. If
    # either of these exceptions is raised, the code redirects the user to the "signIn" route with a
    # message indicating either "Session Kadaluarsa" or "Anda telah logout" (depending on the specific
    # exception caught). This is likely part of a mechanism to handle expired or invalid JWT tokens
    # during user authentication or authorization processes.
    except jwt.ExpiredSignatureError:
        return redirect(url_for("signIn", msg="Session Kadaluarsa"))
    except jwt.DecodeError:
        return redirect(url_for("signIn", msg="Anda telah logout"))
    except Exception as e:
        return (
            redirect(
                url_for(
                    "dashboard",
                    msg=e.args[0] if e.args else "An unexpected error occurred",
                )
            ),
            402,
        )


@app.route("/task/<path>", methods=["POST"])
def task_post_admin(path):
    """
    Manajement Tugas Karyawan / Magang
    ---
    tags:
      - Admin / Sub Admin
    summary: Buat atau Perbarui Tugas Melalui Admin/Sub Admin
    description: Endpoint ini digunakan untuk membuat atau memperbarui tugas untuk pengguna. Memerlukan cookie autentikasi dan token CSRF yang valid.
    parameters:
      - name: path
        in: path
        required: true
        enum: ['add', 'edit', 'delete']
        description: Operasi yang akan dilakukan (misalnya, "add", "edit", "delete").
        type: string
      - name: body
        in: body
        required: true
        description: data json yang diperlukan wajib diisi sesuai keperluan di model.
        schema:
          type: object
          properties:
            taskName:
              type: string
              description: judul task (add)
              example: "judul-task"
            description_task:
              type: string
              description: descripsi task (add)
              example: 'ini deskripsi'
            link_input:
              type: string
              description: link task [https | http] (add)
              example: "https://"
            deadline:
              description: deadline task (add)
              type: string
              format: date-time
              example: "2024-12-25T12:30:00"
            send_to_user:
              type: string
              description: email user Karyawan / Magang (add)
              example: 'xxx@gmail.com'
            jobs:
              type: string
              description: jobs user Karyawan / Magang (add)
              example: 'Karyawan'
            departement:
              type: string
              description: departement user Karyawan / Magang (add)
              example: 'IT'
            rowId_receive:
              type: string
              description: Id user Karyawan / Magang (edit)
              example: 'input-id-user'
            inputId_receive:
              type: string
              description: key task Karyawan / Magang (edit)
              example: 'input-key-user'
            newValue_receive:
              type: string
              description: value baru task Karyawan / Magang (edit)
              example: 'input-value-user'
            'id':
              type: string
              description: id user Karyawan / Magang (delete)
              example: 'input-id-user'
      - name: X-CSRF-Token
        in: header
        required: true
        type: string
        description: masukkan csrf mu
        example: 'your-csrf-form'
      - name: Content-Type
        in: header
        required: true
        type: string
        enum: ['application/json']
        description: your content type
        example: 'application/json'

    responses:
      200:
        description: Tugas berhasil dibuat atau diperbarui.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "Tugas berhasil dibuat"
            status:
              type: string
              example: "sukses"
      400:
        description: Permintaan tidak valid atau parameter yang diperlukan hilang.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "Permintaan tidak valid"
            status:
              type: string
              example: "error"
      401:
        description: Token CSRF atau cookie tidak valid atau telah kedaluwarsa.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "Token CSRF kedaluwarsa"
            status:
              type: string
              example: "error"
      500:
        description: Terjadi kesalahan di server.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "Terjadi kesalahan saat memproses permintaan"
            status:
              type: string
              example: "error"
    """

    try:
        # The above Python code snippet is checking for the presence and validity of a CSRF token and a
        # cookie in a web request. Here is a breakdown of the code:
        cookie = request.cookies.get("token_key")
        csrf_token = request.cookies.get("csrf_token")
        if csrf_token == None:
            return redirect(url_for("signIn", msg="csrf token Kadaluarsa"))
        if cookie == None:
            return redirect(url_for("signIn", msg="Anda Telah logout"))

        # decode uuid
        cookie = uuid_like_to_string(cookie)
        csrf_token = uuid_like_to_string(csrf_token)
        if not csrf_token:
            return redirect(url_for("signIn", msg="CSRF Token Kadaluarsa"))
        if not cookie:
            return redirect(url_for("signIn", msg="Cookie Kadaluarsa"))

        # decode payload
        payload = jwt.decode(cookie, secretKey, algorithms=["HS256"])

        if payload["role"] != 1 and payload["jobs"] not in ("Admin", "Sub Admin"):
            raise Exception("Anda tidak memiliki hak akses")

        # ambil data dari db
        data = db.users.find_one(
            {
                "_id": ObjectId(payload["_id"]),
                "jobs": payload["jobs"],
                "role": payload["role"],
            }
        )

        if not data:
            raise Exception("Anda tidak memiliki hak akses")

        if path == "add":
            # ambil datajsonnnya
            data_json = request.json

            # cek datajosn yang bintang
            for key, value in data_json.items():
                if key != "link_input" and not data_json[key]:
                    raise Exception("form {} wajib diisi".format(key))

            # ambil result id yang bukan role 1 sesuai dengan email json
            result = db.users.find_one(
                {
                    "email": data_json["send_to_user"],
                    "role": {"$ne": 1},
                    "jobs": data_json["jobs"],
                    "departement": data_json["departement"],
                },
                {"_id": 1},
            )

            # result didapatakan?
            if not result:
                print("Terjadi kesalahan dalam pengambilan user di database")
                raise Exception("Terjadi kesalahan pada database")

            data_json["deadline"] = datetime.datetime.strptime(
                data_json["deadline"], "%Y-%m-%dT%H:%M"
            )

            # melakukan penambahan
            result = db.tasks.insert_one(
                {
                    key: value
                    for key, value in data_json.items()
                    if key not in ("send_to_user", "jobs", "departement")
                }
                | {
                    "user_id": result["_id"],
                    "accepted": False,
                    "status_task": "Pending",
                }
            )

            # cek berhasil ditambah?
            if not result:
                print("terjadi kesalahan data dalam penambahan data ditabel task")
                raise Exception("Terjadi kesalahan pada database")

            task_id_new = result.inserted_id

            # kirim ke class task gmail notif
            TaskGmailNotif(task_id_new, path)

            # buat response
            response = make_response(
                jsonify(
                    {
                        "redirect": url_for(
                            "task", status="success", msg="Task Berhasil dibuat"
                        )
                    }
                ),
                200,
            )

            # tambahkan headers
            response.headers["CSRF_TOKEN"] = csrf_token
            response.headers["Content-Type"] = "Application/json"
            return response

        # edit
        elif path == "edit":
            # ambil data json
            rowId, inputId, newValue = request.json.values()

            # cek nilai json tak kosong
            if not rowId and not inputId and not newValue:
                raise Exception("form wajib diisi")

            # decode row id
            rowId = uuid_like_to_string(str(rowId))
            if not rowId:
                raise Exception("Decode Error")

            # new value bila str True and False
            if newValue in ("True", "False"):
                newValue = True if newValue == "True" else False

            if is_valid_datetime_format(newValue):
                newValue = datetime.datetime.strptime(newValue, "%Y-%m-%dT%H:%M")

            # edit database berdasarkan id
            result = db.tasks.find_one_and_update(
                {"_id": ObjectId(rowId)}, {"$set": {inputId: newValue}}
            )

            # cek result
            if not result:
                raise Exception("Terjadi kesalahan pada database")

            # ambil task id
            task_id_new = result["_id"]

            # kirim ke class task gmail notif
            TaskGmailNotif(task_id_new, path, newValue=newValue)

            # buat response
            return make_response(
                jsonify({"redirect": url_for("task", msg="Task Berhasil diubah")}), 200
            )

        elif path == "delete":
            # ambil data json
            rowId = request.json.get("id")
            rowId = uuid_like_to_string(str(rowId))

            # cek nilai json tak kosong
            if not rowId:
                raise Exception("form wajib diisi")

            # edit database berdasarkan id
            result = db.tasks.delete_one({"_id": ObjectId(rowId)})

            # cek result
            if not result:
                raise Exception("Terjadi kesalahan pada database")

            # buat response
            return make_response(
                jsonify({"redirect": True}),
                200,
            )

        else:
            raise Exception("Path doesn`t exists")
    # The above code is handling exceptions related to JWT (JSON Web Token) in a Python application.
    # Specifically, it is catching `jwt.ExpiredSignatureError` and `jwt.DecodeError` exceptions. If
    # either of these exceptions is raised, the code redirects the user to the "signIn" route with a
    # message indicating either "Session Kadaluarsa" or "Anda telah logout" (depending on the specific
    # exception caught). This is likely part of a mechanism to handle expired or invalid JWT tokens
    # during user authentication or authorization processes.
    except jwt.ExpiredSignatureError:
        return make_response(
            jsonify({"redirect": url_for("signIn", msg="Session Kadaluarsa")}), 500
        )
    except jwt.DecodeError:
        return make_response(
            jsonify({"redirect": url_for("signIn", msg="Anda telah logout")}), 500
        )
    except Exception as e:
        # jikakosong
        if not e.args:
            return make_response(
                jsonify(
                    {
                        "redirect": url_for(
                            "dashboard", msg="An unexpected error occurred"
                        )
                    }
                ),
                500,
            )
        return make_response(jsonify({"redirect": url_for("task", msg=e.args[0])}), 500)


@app.route("/task/user/<path>", methods=["POST"])
def task_post_user(path):
    """
    Manajement Progress Tugas
    ---
    tags:
      - Magang / Karyawan
    summary: Buat atau Perbarui Progress Tugas Pengguna
    description: Endpoint ini digunakan untuk membuat atau memperbarui progress tugas pengguna. Memerlukan cookie autentikasi dan token CSRF yang valid untuk mengakses data.
    parameters:
      - name: path
        in: path
        required: true
        enum: ['edit']
        description: Operasi yang akan dilakukan (misalnya, "edit").
        type: string

      - name: body
        in: body
        required: true
        description: wajib diisi sesuai keperluan di model.
        schema:
          type: object
          properties:
            parenTrId_receive:
              type: string
              description: id task yang ingin diupdate progressnya (edit)
              example: "input-id-task"
            newValue_receive:
              type: string
              description: value baru update task (edit)
              example: 'input-value-user'

      - name: X-CSRF-Token
        in: header
        required: true
        type: string
        description: masukkan csrf mu
        example: 'your-csrf-form'

      - name: Content-Type
        in: header
        required: true
        type: string
        enum: ['application/json']
        description: your content type
        example: 'application/json'

    responses:
      200:
        description: Tugas berhasil dibuat atau diperbarui.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "Tugas berhasil diperbarui"
            status:
              type: string
              example: "sukses"
      400:
        description: Permintaan tidak valid atau parameter yang diperlukan hilang.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "Permintaan tidak valid"
            status:
              type: string
              example: "error"
      401:
        description: Token CSRF atau cookie tidak valid atau telah kedaluwarsa.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "Token CSRF kedaluwarsa"
            status:
              type: string
              example: "error"
      500:
        description: Terjadi kesalahan di server.
        schema:
          type: object
          properties:
            message:
              type: string
              example: "Terjadi kesalahan saat memproses permintaan"
            status:
              type: string
              example: "error"
    """

    try:
        # The above Python code snippet is checking for the presence and validity of a CSRF token and a
        # cookie in a web request. Here is a breakdown of the code:
        cookie = request.cookies.get("token_key")
        csrf_token = request.cookies.get("csrf_token")
        if csrf_token == None:
            return redirect(url_for("signIn", msg="csrf token Kadaluarsa"))
        if cookie == None:
            return redirect(url_for("signIn", msg="Anda Telah logout"))

        # decode uuid
        cookie = uuid_like_to_string(cookie)
        csrf_token = uuid_like_to_string(csrf_token)
        if not csrf_token:
            return redirect(url_for("signIn", msg="CSRF Token Kadaluarsa"))
        if not cookie:
            return redirect(url_for("signIn", msg="Cookie Kadaluarsa"))

        # decode payload
        payload = jwt.decode(cookie, secretKey, algorithms=["HS256"])

        if payload["role"] == 1 and payload["jobs"] not in ("Karyawan", "Magang"):
            raise Exception("Anda tidak memiliki hak akses")

        data = db.users.find_one(
            {
                "_id": ObjectId(payload["_id"]),
                "jobs": payload["jobs"],
                "role": payload["role"],
            }
        )

        if not data:
            raise Exception("Anda tidak memiliki hak akses")

        if path == "edit":
            # The above Python code is checking if the `X-CSRFToken` and `Content-Type` headers are
            # present in the request headers. If either of these headers is missing (`csrf_token` or
            # `contents` is `None`), the code will execute the block inside the `if` statement.
            csrf_token = request.headers.get("X-CSRFToken")
            contents = request.headers.get("Content-Type")
            if not csrf_token or not contents:
                raise Exception("anda melakukan ajax diluar")

            # The code snippet is written in Python and it is extracting the values of `task_id` and
            # `status_task_new` from a JSON request. It then checks if either `task_id` or
            # `status_task_new` is empty (evaluates to False).
            task_id, status_task_new = request.json.values()
            if not task_id or not status_task_new:
                raise Exception("Ada kesalahan dalam meload data task")

            # The code snippet is attempting to convert a UUID-like object to a string representation
            # using the `uuid_like_to_string` function. If the conversion is successful (i.e., the
            # resulting string is not empty), the code proceeds with the next steps.
            task_id_conv = uuid_like_to_string(str(task_id))
            if not task_id_conv:
                raise Exception("Terjadi kesalahan dalam mengonvert data task")

            # The above Python code is querying a MongoDB database to find a specific task document
            # based on the user ID and task ID provided. It retrieves the '_id' and 'status_task'
            # fields of the found document. If no document is found (status_task_old is None), it
            # prints an error message indicating that there was an issue retrieving the status of the
            # old task from the database.
            status_task_old = db.tasks.find_one(
                {"user_id": data["_id"], "_id": ObjectId(task_id_conv)},
                {"_id": 1, "status_task": 1},
            )
            if not status_task_old:
                print("Terjadi kesalahan dalam pengambil status task old di db")
                raise Exception("Terjadi kesalahan di db")

            # The code is checking if the value of the 'status_task' key in the dictionary
            # `status_task_old` is equal to the value of the variable `status_task_new`. If they are
            # equal, the condition will be true.
            if status_task_old["status_task"] == status_task_new:
                raise Exception("Tidak ada perubahan status task")

            # The above code is updating a document in a MongoDB collection called "tasks". It is
            # finding a document with a specific "_id" and "user_id" that matches the provided values,
            # and then updating the "status_task" field with a new value specified by
            # "status_task_new". If the update is successful, the result will contain information
            # about the update operation.
            result = db.tasks.update_one(
                {"_id": ObjectId(task_id_conv), "user_id": data["_id"]},
                {"$set": {"status_task": status_task_new}},
            )
            if not result:
                raise Exception("Terjadi kesalahan dalam update db")

            # The above code is creating a response with a JSON object that includes a redirect URL
            # for a specific task with a success status message. The response is returned with a
            # status code of 200 (OK).
            return make_response(
                jsonify(
                    {
                        "redirect": url_for(
                            "task", status="success", msg="status berhasil di update"
                        )
                    }
                ),
                200,
            )
        else:
            raise Exception("Path not found! Change this!")

    # The above code is handling exceptions related to JWT (JSON Web Token) in a Python application.
    # Specifically, it is catching `jwt.ExpiredSignatureError` and `jwt.DecodeError` exceptions. If
    # either of these exceptions is raised, the code redirects the user to the "signIn" route with a
    # message indicating either "Session Kadaluarsa" or "Anda telah logout" (depending on the specific
    # exception caught). This is likely part of a mechanism to handle expired or invalid JWT tokens
    # during user authentication or authorization processes.
    except jwt.ExpiredSignatureError:
        return make_response(
            jsonify({"redirect": url_for("signIn", msg="Session Kadaluarsa")}), 500
        )
    except jwt.DecodeError:
        return make_response(
            jsonify({"redirect": url_for("signIn", msg="Anda telah logout")}), 500
        )
    except Exception as e:
        # jikakosong
        if not e.args:
            return make_response(
                jsonify(
                    {
                        "redirect": url_for(
                            "dashboard", msg="An unexpected error occurred"
                        )
                    }
                ),
                500,
            )
        return make_response(jsonify({"redirect": url_for("task", msg=e.args[0])}), 500)


# melakukan error handler csrf
@app.route("/400", methods=["POST"])
@app.errorhandler(400)
def handle_csrf_error(e=None):
    """
    Handle CSRF error
    ---
    tags:
      - Error
    summary: Tangani Kesalahan Token CSRF
    description: Endpoint ini menangani kesalahan akibat token CSRF tidak valid atau permintaan tidak valid.
    responses:
      "400":
        description: Token CSRF tidak valid.
        content:
          application/json:
            schema:
              type: object
              properties:
                message:
                  type: string
                  example: Token CSRF tidak valid atau kedaluwarsa
                status:
                  type: string
    """
    # Ambil pesan error sebagai string (atau fallback jika None)
    error_message = str(e) if e else "Permintaan tidak valid"
    return jsonify({"error": error_message}), 400


@app.route("/404", methods=["GET"])
@app.errorhandler(404)
def notFound(error=None):
    """
    Halaman Tidak Ditemukan
    ---
    tags:
      - Error
    summary: Halaman Tidak Ditemukan
    description: Menampilkan halaman kesalahan ketika pengguna mengakses URL yang tidak ada.
    parameters:
      - name: error
        in: query
        required: false
        type: string
        description: Parameter opsional untuk pesan error.
    responses:
      "200":
        description: Halaman ditemukan.
        content:
          application/json:
            example:
              message: Halaman ditemukan.
              status: ok
      "404":
        description: Halaman tidak ditemukan.
        content:
          application/json:
            example:
              message: Halaman yang Anda cari tidak ditemukan.
              status: error
    """
    print("error", error)
    data = {"next": "/", "previous": "javascript:history.back()"}
    return render_template("notFound.html", data=data), 200


@app.route("/api/cron_task", methods=["GET"])
def cron_task():
    """
    API untuk menghandle cron job pengecekan absensi
    ---
    tags:
      - Cron Job
    summary: API untuk menghandle cron job pengecekan absensi
    description: API ini digunakan untuk menghandle cron job pengecekan absensi setiap menit.
    responses:
      "200":
        description: Berhasil menghandle cron job pengecekan absensi.
        content:
          application/json:
            example:
              message: success
              data: []
      "500":
        description: Terjadi kesalahan saat menghandle cron job pengecekan absensi.
        content:
          application/json:
            example:
              message: Terjadi kesalahan saat menghandle cron job pengecekan absensi.
    """
    # try:
    response = unhadir_absensi()
    return make_response(jsonify({"message": "success", "data": response}), 200)
    # except Exception as e:
    #     return make_response(jsonify({"message": str(e)}), 500)


# stating app
if __name__ == "__main__":
    # # Menjadwalkan pengecekan absensi setiap menit
    # delete_absen = BackgroundScheduler()
    # delete_absen.add_job(
    #     func=unhadir_absensi, trigger="interval", minutes=1
    # )  # interval hours/minute/second. date run_date .cron day_of_week,hours,minutes
    # delete_absen.start()
    app.run(port=8080, debug=True)  # ssl_context =  adhoc adalah sertifikat self signed
    # DEBUG is SET to TRUE. CHANGE FOR PROD
