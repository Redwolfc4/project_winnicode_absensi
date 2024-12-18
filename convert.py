from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.drawing.image import Image
import os
from fpdf import FPDF
from urllib.parse import math

# convert to pdf


class PDF(FPDF):
    """
    Class ini digunakan untuk membuat pdf dari data excel
    """

    def create_pdf(
        self, start_row: int, end_row: int, column_widths: list, ws, currentPage=None
    ):
        from app import app

        """
        Fungsi ini digunakan untuk membuat pdf dari data excel
        parameter:
            start_row (int): baris awal data yang akan di export
            end_row (int): baris akhir data yang akan di export
            column_widths (list): list berisi lebar kolom yang telah di sesuaikan
            ws (Worksheet): worksheet yang akan di export
        return:
            None|fpdf

        Contoh pemakaian:
            pdf = PDF()
            pdf.create_pdf(start_row=1,end_row=10,column_widths=[10,20,30,40],ws=ws)
        """

        # The above code snippet in Python is initializing instance variables in a class constructor.
        # It assigns values to private variables `__ws`, `__start_row`, `__end_row`,
        # `__column_widths`, `__margin_x`, and `__page_width`. The `__margin_x` is set to 2.3, and
        # `__page_width` is calculated as the ceiling of the difference between `self.w` and `2.3`
        # multiplied by 2.
        self.__ws = ws
        self.__start_row = start_row
        self.__end_row = end_row
        self.__column_widths = column_widths
        self.__margin_x = 2.3
        self.__page_width = math.ceil(self.w - self.__margin_x * 2)

        # tambahkan judul
        self.set_font("Times", "B", 14)
        self.set_text_color(255, 102, 196)
        if currentPage == "Kelola Admin":
            self.cell(self.__page_width - 10, 10, "Data Admin / Sub Admin", 0, 1, "C")
        else:
            self.cell(self.__page_width, 10, "Data Karyawan / Magang", 0, 1, "C")
        self.cell(
            (
                self.__page_width - 10
                if currentPage == "Kelola Admin"
                else self.__page_width
            ),
            10,
            "PT. Winnicode Garuda Teknologi",
            0,
            1,
            "C",
        )
        if currentPage == "Kelola Admin":
            self.set_margins(self.__margin_x + 55, self.__margin_x, self.__margin_x)
        else:
            self.set_margins(self.__margin_x, self.__margin_x, self.__margin_x)
        self.ln(10)

        # buat table heading
        self.set_font("Times", "B", 6.5)
        self.set_text_color(0, 0, 0)
        self.set_fill_color(255, 102, 196)
        for row in self.__ws.iter_rows(
            min_row=self.__start_row, max_row=self.__start_row
        ):
            for header_cell in row:
                column = (
                    header_cell.column_letter
                )  # Mendapatkan huruf kolom (A, B, C, dst.)
                header_cell_value = (
                    str(header_cell.value) if header_cell.value else ""
                )  # Nilai sel sebagai string
                self.cell(
                    self.__column_widths[column] + 3.1,
                    5,
                    header_cell_value,
                    1,
                    0,
                    "C",
                    fill=True,
                )
        self.ln()

        # Set font untuk tabel data
        self.set_font("Times", "", 6.5)

        # The above Python code is a loop that iterates over rows in a worksheet and processes each
        # cell within the row. Here is a summary of what the code is doing:
        # enter data
        for row in self.__ws.iter_rows(
            min_row=self.__start_row + 1, max_row=self.__end_row
        ):
            max_lines = 0
            for data_cell in row:
                column = (
                    data_cell.column_letter
                )  # Mendapatkan huruf kolom (A, B, C, dst.)
                data_cell_value = (
                    str(data_cell.value) if data_cell.value else ""
                )  # Nilai sel sebagai string
                num_lines = (
                    self.get_string_width(str(data_cell.value))
                    / self.__column_widths[column]
                )
                max_lines = max(max_lines, num_lines)

            row_height = max_lines * 10  # Adjust row height based on maximum lines

            # The above code is a Python loop that iterates over each element in the variable `row`
            # and assigns it to the variable `data_cell`.
            # enter data
            for data_cell in row:
                column = (
                    data_cell.column_letter
                )  # Mendapatkan huruf kolom (A, B, C, dst.)
                data_cell_value = (
                    str(data_cell.value).strip() if data_cell.value else ""
                )  # Nilai sel sebagai string

                # The above code is checking if the variable `data_cell_value` is a string and if it
                # ends with either '.png', '.jpg', or '.jpeg'. If both conditions are true, the code
                # block inside the if statement will be executed.
                if isinstance(data_cell_value, str) and data_cell_value.endswith(
                    (".png", ".jpg", ".jpeg")
                ):
                    img_path = os.path.join(app.root_path, "static/") + str(
                        data_cell_value
                    )
                    cell_width = self.__column_widths[column] + 3.1
                    cell_height = 10  # Tinggi sel
                    x = self.get_x()
                    y = self.get_y()
                    # Jika file gambar ditemukan
                    if os.path.exists(img_path.strip()):
                        # Tambahkan border pada sel sebelum gambar dimasukkan
                        self.cell(cell_width, cell_height, border=1)
                        # tambahkan image
                        self.image(img_path, x + 2, y + 1, cell_width - 4, 8)
                        # Geser posisi X untuk sel berikutnya
                        self.set_xy(x + self.__column_widths[column] + 3.1, y)
                        continue

                # If the text is too long, use multi_cell to wrap it
                if (
                    len(data_cell_value.strip()) > self.__column_widths[column]
                ):  # Adjust this limit as needed
                    # Store current position
                    x = self.get_x()
                    y = self.get_y()
                    self.multi_cell(
                        self.__column_widths[column] + 3.1,
                        5,
                        data_cell_value,
                        1,
                        "C",
                        split_only=False,
                    )
                    # # Move to the right for the next cell in the same row
                    self.set_xy(x + self.__column_widths[column] + 3.1, y)
                else:
                    # For shorter texts, use cell (no wrapping needed)
                    self.cell(
                        self.__column_widths[column] + 3.1,
                        10,
                        data_cell_value,
                        1,
                        0,
                        "C",
                    )
            self.ln(10)

        return self


# convert ke excel
def convert_to_excel(ws, result, currentPage=None, start=None, stop=None):
    from app import app

    """
    Converts data to Excel format and writes it to the provided worksheet.

    This function processes the given data `result`, cleans it by removing any
    dictionary values, and appends the cleaned data to the provided worksheet `ws`.
    It also sets up borders and adjusts column widths and row heights for better
    readability in the Excel file.

    Parameters:
    ws (Worksheet): The worksheet where the data will be written.
    result (list): A list of dictionaries containing the data to be converted.
    start (int, optional): The starting index for processing the `result` list. Defaults to None.
    stop (int, optional): The stopping index for processing the `result` list. Defaults to None.

    Returns:
    None

    Example:
    >>> from openpyxl import Workbook
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> result = [
    ...     {"name": "John Doe", "age": 30, "absen": {"hadir": 10, "tidak_hadir": 2}},
    ...     {"name": "Jane Doe", "age": 25, "absen": {"hadir": 12, "tidak_hadir": 0}},
    ... ]
    >>> convert_to_excel(ws, result)
    >>> wb.save("output.xlsx")
    """

    thin = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    column_widths = {}
    row_heights = 0

    # data tanpa dictionary
    for i, x in enumerate(result):
        # Convert dict_values to list and remove dictionaries
        cleaned_data = [item for item in x.values() if not isinstance(item, dict)]
        cleaned_data.insert(0, i + 1)

        if currentPage != "Kelola Admin":
            cleaned_data.append(x["absen"]["hadir"])
            cleaned_data.append(x["absen"]["telat"])
            cleaned_data.append(x["absen"]["tidak_hadir"])
            cleaned_data.append(x["absen"]["libur"])

        ws.append(cleaned_data)  # menambhkan ke ws
        row = ws.max_row  # mencari row mana yang ditambah

        if start == None and stop == None:
            start = row
            stop = row
        else:
            if row > stop:
                stop = row
            if row < start:
                start = row

        for col in range(1, len(cleaned_data) + 1):  # dari 1 sampai banyak database
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).font = Font(name="Times", size=11)
            ws.cell(row=row, column=col).alignment = Alignment(
                horizontal="center", vertical="center"
            )

    # iterasi width
    for row in ws.iter_rows(min_row=start - 1, max_row=stop):
        for cell in row:
            column = cell.column_letter  # Mendapatkan huruf kolom (A, B, C, dst.)
            cell_value = (
                str(cell.value).strip() if cell.value else ""
            )  # Nilai sel sebagai string

            if isinstance(cell_value, str) and cell_value.endswith(
                (".png", ".jpg", ".jpeg")
            ):
                img_path = os.path.join(app.root_path, "static/") + str(cell_value)

                # Jika file gambar ditemukan
                if os.path.exists(img_path.strip()):
                    # Load gambar dan masukkan ke cell
                    cell.number_format = ";;;"
                    img = Image(img_path.strip())
                    img_width = img.width = 50
                    img_height = img.height = 50  # Tinggi gambar dalam piksel
                    img.anchor = f"{column}{cell.row}"
                    img.bottom = 100

                    # Konversi ukuran gambar ke ukuran kolom dan baris
                    # 1 kolom Excel ≈ 7 pixel, 1 row Excel ≈ 0.75 point
                    column_width = img_width / 7
                    row_height = img_height / 1.12

                    # Kosongkan nilai teks di cell setelah gambar dimasukkan
                    if column in column_widths:
                        column_widths[column] = max(column_widths[column], column_width)
                    else:
                        column_widths[column] = column_width + 2

                    row_heights = max(row_heights, row_height)

                    ws.add_image(img)
                    continue

            # Simpan panjang maksimum di setiap kolom
            if column in column_widths:
                column_widths[column] = max(column_widths[column], len(cell_value))
            else:
                column_widths[column] = len(cell_value)

    # Sesuaikan lebar kolom berdasarkan panjang maksimum
    for col, width in column_widths.items():
        # Tambahkan padding agar tidak terlalu sempit
        ws.column_dimensions[col].width = width + 2

    for row in range(start, stop + 1):
        ws.row_dimensions[row].height = int(row_heights)
    return (column_widths, start, stop)
